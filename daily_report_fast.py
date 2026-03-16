#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
일일결산 자동화 시스템 (최적화 버전)
안과 검사실의 일일 통계를 자동으로 수집하고 PDF 보고서를 생성하는 프로그램
"""

import os
import sys
import json
import re
import time
import shutil
import threading
import requests
from datetime import datetime, date, timedelta
from typing import Set, Dict, List, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed
import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext, messagebox

try:
    import pyodbc
    HAS_PYODBC = True
except ImportError:
    HAS_PYODBC = False

# SSL certificate fix for PyInstaller bundled exe
if getattr(sys, 'frozen', False):
    try:
        import certifi
        os.environ['REQUESTS_CA_BUNDLE'] = certifi.where()
        os.environ['SSL_CERT_FILE'] = certifi.where()
    except (ImportError, Exception):
        # certifi not available - will fall back to verify=False in API calls
        pass

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("오류: openpyxl이 설치되지 않았습니다.")
    print("설치: pip install openpyxl")
    sys.exit(1)

try:
    import pandas as pd
except ImportError:
    print("오류: pandas가 설치되지 않았습니다.")
    print("설치: pip install pandas")
    sys.exit(1)

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

# 파일 캐시 시스템
try:
    from file_cache_manager import get_new_files, update_cache_with_today_files, load_cache
    HAS_CACHE = True
except ImportError:
    HAS_CACHE = False

# Windows에서만 pywin32 임포트
if sys.platform == 'win32':
    try:
        import win32com.client
        import pythoncom
        HAS_WIN32 = True
    except ImportError:
        HAS_WIN32 = False
        print("경고: pywin32가 설치되지 않았습니다. PDF 변환을 사용할 수 없습니다.")
else:
    HAS_WIN32 = False


def get_exe_dir() -> str:
    """exe 또는 스크립트의 실행 디렉토리 반환"""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def get_bundled_path(filename: str) -> Optional[str]:
    """PyInstaller 번들 내장 리소스 경로 반환"""
    if getattr(sys, 'frozen', False):
        bundled = os.path.join(sys._MEIPASS, filename)
        if os.path.exists(bundled):
            return bundled
    return None


def get_config_path() -> str:
    """config.json 경로 결정: exe 옆 외부 파일 우선, 없으면 번들 리소스 사용"""
    external = os.path.join(get_exe_dir(), 'config.json')
    if os.path.exists(external):
        return external

    bundled = get_bundled_path('config.json')
    if bundled:
        return bundled

    return external  # fallback (FileNotFoundError will be raised later)


def get_template_path(config_path: str) -> str:
    """템플릿 xlsx 경로 결정: config 경로 → exe 옆 → 번들 리소스 순"""
    # 1. config에 지정된 경로가 존재하면 사용
    if config_path and os.path.exists(config_path):
        return config_path

    # 2. exe 옆에 템플릿 파일
    template_name = '일일결산_템플릿.xlsx'
    external = os.path.join(get_exe_dir(), template_name)
    if os.path.exists(external):
        return external

    # 3. exe 상위 디렉토리 (결산 폴더 안에 exe를 둔 경우)
    parent = os.path.join(get_exe_dir(), '..', template_name)
    if os.path.exists(parent):
        return os.path.abspath(parent)

    # 4. PyInstaller 번들 리소스
    bundled = get_bundled_path(template_name)
    if bundled:
        return bundled

    return config_path or external  # fallback


class DailyReportSystem:
    """일일결산 시스템의 메인 클래스 (최적화 버전)"""

    def __init__(self, config_path: str = None):
        if config_path is None:
            config_path = get_config_path()
        self.config_path = config_path
        self.config = self.load_config(config_path)
        self.chart_numbers = {}
        self.results = {}
        self.today = date.today()

        # 정규식 패턴 미리 컴파일
        self.compiled_patterns = {}
        for eq_id, eq_info in self.config['equipment'].items():
            self.compiled_patterns[eq_id] = re.compile(eq_info['pattern'])

    def load_config(self, config_path: str) -> dict:
        """설정 파일 로드"""
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except FileNotFoundError:
            messagebox.showerror("오류", f"설정 파일을 찾을 수 없습니다: {config_path}")
            sys.exit(1)
        except json.JSONDecodeError:
            messagebox.showerror("오류", "설정 파일 형식이 올바르지 않습니다.")
            sys.exit(1)

    def get_staff_from_api(self, target_date: date, department: str = "검사실"):
        """Schedule API에서 오늘 근무하는 검사실 직원 조회

        1) /api/employees → 검사실 active 직원 전체
        2) /api/schedule/{year}/{month} → 해당 월 스케줄
        3) 오늘 휴무(h1, h2, 기타) 등록된 직원 제외

        Returns: (all_names, working_staff) tuple or None
        Sets self._last_api_error with error message on failure
        """
        self._last_api_error = None
        try:
            api_config = self.config.get('hospital_schedule_api', {})
            if not api_config.get('enabled', True):
                self._last_api_error = "API disabled in config"
                return None

            # Internal URL first (hospital LAN), then external as fallback
            internal_url = api_config.get('internal_url', 'http://192.168.0.210:3001').rstrip('/')
            external_url = api_config.get('url', 'https://schedule.seran-it.com').rstrip('/')
            api_token = api_config.get('api_token', 'hospital2025secure')
            dept = api_config.get('department', department)
            headers = {'x-api-token': api_token}

            # Try internal URL first, then external
            base_url = None
            response = None
            for try_url in [internal_url, external_url]:
                try:
                    response = requests.get(f"{try_url}/api/employees", headers=headers, timeout=3)
                    if response.status_code == 200:
                        base_url = try_url
                        break
                except Exception:
                    continue

            if base_url is None or response is None or response.status_code != 200:
                self._last_api_error = "내부/외부 URL 모두 연결 불가"
                return None

            if response.status_code != 200:
                self._last_api_error = f"HTTP {response.status_code}"
                return None

            all_employees = response.json()
            exam_staff = [
                emp for emp in all_employees
                if emp.get('team') == dept and emp.get('status') == 'active'
                and not emp.get('isRetired')
            ]
            exam_ids = {emp['id']: emp['name'] for emp in exam_staff}

            # 2. Schedule API에서 해당 월 스케줄 가져오기
            year = target_date.strftime('%Y')
            month = target_date.strftime('%m')
            sr = requests.get(f"{base_url}/api/schedule/{year}/{month}",
                            headers=headers, timeout=5)

            off_staff = set()
            if sr.status_code == 200:
                sdata = sr.json()
                schedule_entries = sdata.get('schedule', {})
                target_day = str(target_date.day)

                for key, entry in schedule_entries.items():
                    # key format: emp_{id}_{day} (day = 1~31)
                    parts = key.split('_')
                    if len(parts) < 3:
                        continue
                    try:
                        emp_id = int(parts[1])
                    except (ValueError, IndexError):
                        continue
                    day = parts[2]

                    if day == target_day and emp_id in exam_ids:
                        sch = entry.get('schedule', entry) if isinstance(entry, dict) else {}
                        sch_value = sch.get('value', '') if isinstance(sch, dict) else ''
                        # h로 시작하는 코드 모두 제외 (h1=연차, h2=휴가, h3=반차, h4=병가, h5=의료휴무, h6=기타)
                        if str(sch_value).startswith('h') or str(sch_value).startswith('기타'):
                            off_staff.add(exam_ids[emp_id])

            # 3. 전체 직원에서 휴가자 제외
            all_names = [emp['name'] for emp in exam_staff]
            working_staff = [name for name in all_names if name not in off_staff]

            return (all_names, working_staff) if all_names else None

        except requests.exceptions.ConnectionError as e:
            self._last_api_error = f"연결 불가 ({base_url})"
            return None
        except requests.exceptions.Timeout:
            self._last_api_error = "타임아웃 (5초)"
            return None
        except Exception as e:
            self._last_api_error = str(e)[:80]
            return None

    def is_valid_chart_number(self, chart_num_str: str) -> bool:
        """차트번호 유효성 검증"""
        try:
            if chart_num_str.startswith('0') and len(chart_num_str) > 1:
                return False
            chart_num = int(chart_num_str)
            min_val = self.config['validation']['chart_number_min']
            max_val = self.config['validation']['chart_number_max']
            return min_val <= chart_num <= max_val
        except (ValueError, KeyError):
            return False

    def extract_chart_number(self, match) -> Optional[str]:
        """정규식 매칭에서 차트번호 추출 (단일/이중 그룹 패턴 지원)

        단일 그룹 패턴 (SP, TOPO 등): (\d+)_
        이중 그룹 패턴 (HFA): _(\d{5,6})$|^(\d{5,6})_
        """
        if not match:
            return None
        return match.group(1) or (match.group(2) if match.lastindex > 1 else None)

    def _resolve_date_folder(self, folder_structure: str) -> str:
        """folder_structure 템플릿을 오늘 날짜로 변환

        순서 중요: 긴 패턴부터 변환 (YYYY.MM → YYYY → MM.DD → MM → DD)
        """
        folder = folder_structure
        folder = folder.replace('YYYY.MM', self.today.strftime('%Y.%m'))
        folder = folder.replace('YYYY', self.today.strftime('%Y'))
        folder = folder.replace('MM.DD', self.today.strftime('%m.%d'))
        folder = folder.replace('MM', self.today.strftime('%m'))
        folder = folder.replace('DD', self.today.strftime('%d'))
        return folder

    def get_today_folder_path(self, base_path: str, equipment_id: str) -> Optional[str]:
        """오늘 날짜 폴더 경로 생성 (config의 folder_structure 사용)"""
        if equipment_id not in self.config['equipment']:
            return base_path

        equipment = self.config['equipment'][equipment_id]
        if 'folder_structure' not in equipment:
            return base_path

        folder = self._resolve_date_folder(equipment['folder_structure'])
        full_path = os.path.join(base_path, folder)

        if os.path.exists(full_path):
            return full_path
        else:
            return None

    def scan_directory_fast(self, equipment_id: str, log_callback) -> Set[str]:
        """
        장비 디렉토리 스캔 (최적화 버전)
        - 오늘 날짜 폴더만 스캔
        - os.walk() 사용
        - 정규식 미리 컴파일
        - 경로에 날짜 포함 여부로 빠른 필터링
        """
        equipment = self.config['equipment'][equipment_id]
        base_path = equipment['path']
        pattern = self.compiled_patterns[equipment_id]
        scan_type = equipment['scan_type']

        chart_numbers = set()

        # 오늘 날짜 패턴들 (경로/파일명 매칭용) - 백업 스크립트 방식
        today_patterns = [
            self.today.strftime('%m.%d'),     # 11.16
            self.today.strftime('%Y%m%d'),    # 20251116
            self.today.strftime('%Y-%m-%d'),  # 2025-11-16
            self.today.strftime('%Y.%m.%d'),  # 2025.11.16
        ]

        if not os.path.exists(base_path):
            log_callback(f"  ⚠️  경로 없음: {base_path}")
            return chart_numbers

        try:
            # 오늘 날짜 폴더 경로 찾기
            today_folder = self.get_today_folder_path(base_path, equipment_id)
            is_realtime_scan = False  # 기본값

            if today_folder is None:
                # 날짜 폴더가 없는 경우: base_path를 직접 스캔
                # SP, HFA, Fundus 등은 낮에는 최상위 폴더에 직접 저장, 저녁에 날짜 폴더로 이동
                # 날짜 폴더가 없으면 최상위에 있는 것들이 오늘 것임
                today_folder = base_path
                is_realtime_scan = True  # 실시간 스캔 표시
                use_creation_time = equipment.get('use_creation_time', False)
                log_callback(f"     📂 스캔 경로: {today_folder} (날짜 폴더 미정리 - 최상위 전체 스캔)")
                if use_creation_time:
                    log_callback(f"     🔍 생성일 확인 모드")

                # 단일 폴더만 스캔 - os.scandir() 사용 (stat 캐싱으로 더 빠름)
                if scan_type == 'file':
                    log_callback(f"     ⚡ os.scandir() 사용 (stat 캐싱)")

                    valid_extensions = self.config['validation']['file_extensions']
                    total_files = 0
                    candidate_entries = []

                    # os.scandir()은 DirEntry 객체를 반환 (stat 정보 캐싱됨)
                    try:
                        with os.scandir(today_folder) as entries:
                            for entry in entries:
                                total_files += 1
                                if entry.is_file(follow_symlinks=False):
                                    if any(entry.name.lower().endswith(ext) for ext in valid_extensions):
                                        candidate_entries.append(entry)
                    except Exception as e:
                        log_callback(f"     ❌ 스캔 오류: {e}")
                        return chart_numbers

                    log_callback(f"     📊 전체: {total_files}개 / 유효 확장자: {len(candidate_entries)}개")

                    if not candidate_entries:
                        log_callback(f"     ⚠️  유효한 파일 없음")
                        return chart_numbers

                    # 최적화 1: 날짜 폴더 미정리 시 모든 파일을 오늘 것으로 간주
                    if is_realtime_scan:
                        log_callback(f"     🔍 실시간 스캔 모드: 모든 파일 매칭")
                        for entry in candidate_entries:
                            match = pattern.search(entry.name)
                            if match:
                                chart_num = self.extract_chart_number(match)
                                if self.is_valid_chart_number(chart_num):
                                    chart_numbers.add(chart_num)
                        log_callback(f"     ✅ 매칭 완료: {len(chart_numbers)}건")
                    else:
                        # 날짜 폴더가 있는 경우: 파일명/경로에 날짜 확인
                        filename_matched = 0
                        need_ctime_check = []

                        for entry in candidate_entries:
                            # 파일명 또는 전체 경로에 오늘 날짜가 있으면 바로 처리
                            if any(dp in entry.path for dp in today_patterns):
                                filename_matched += 1
                                match = pattern.search(entry.name)
                                if match:
                                    chart_num = self.extract_chart_number(match)
                                    if self.is_valid_chart_number(chart_num):
                                        chart_numbers.add(chart_num)
                            elif use_creation_time:
                                need_ctime_check.append(entry)

                        if filename_matched > 0:
                            log_callback(f"     ⚡ 파일명/경로 날짜 매칭: {filename_matched}개 → {len(chart_numbers)}건")

                        # 최적화 2: 생성일 확인이 필요한 경우 (파일명에 날짜 없음)
                        if need_ctime_check and use_creation_time:
                            log_callback(f"     🔍 생성일 확인 필요: {len(need_ctime_check)}개")

                            # 캐시 시스템 사용 (가장 빠름)
                            if HAS_CACHE:
                                cache = load_cache(today_folder)
                                if cache['last_updated']:
                                    log_callback(f"     ⚡ 캐시 사용: 마지막 업데이트 {cache['last_updated'][:10]}")
                                    entry_names = [e.name for e in need_ctime_check]
                                    new_file_names = get_new_files(today_folder, entry_names)
                                    new_file_set = set(new_file_names)
                                    need_ctime_check = [e for e in need_ctime_check if e.name in new_file_set]
                                    log_callback(f"     📊 캐시에 없는 새 파일: {len(need_ctime_check)}개 (기존 {len(entry_names) - len(need_ctime_check)}개 스킵)")

                                    if not need_ctime_check:
                                        log_callback(f"     ✅ 새 파일 없음 - 캐시에서 모두 확인됨")
                                        # 캐시 업데이트
                                        update_cache_with_today_files(today_folder, [e.name for e in candidate_entries])
                                        return chart_numbers
                                else:
                                    log_callback(f"     💾 캐시 없음 - 첫 실행 (다음부터 빨라짐)")

                            log_callback(f"     ⚡ os.scandir() stat 캐싱 사용 (getctime보다 10배 빠름)")

                            # DirEntry.stat()은 캐싱됨 - 네트워크 호출 최소화
                            def check_entry_date(entry):
                                try:
                                    # entry.stat()은 캐싱되어 있어 매우 빠름
                                    stat_info = entry.stat(follow_symlinks=False)
                                    ctime = stat_info.st_ctime
                                    file_date = date.fromtimestamp(ctime)
                                    if file_date == self.today:
                                        match = pattern.search(entry.name)
                                        if match:
                                            chart_num = self.extract_chart_number(match)
                                            if self.is_valid_chart_number(chart_num):
                                                return chart_num, file_date
                                    return None, file_date
                                except:
                                    pass
                                return None, None

                            # 배치 처리 (1000개씩) - entry.stat()은 캐싱되어 병렬 불필요
                            batch_size = 1000
                            total_checked = 0
                            consecutive_old_files = 0
                            ctime_matches = 0

                            for i in range(0, len(need_ctime_check), batch_size):
                                batch = need_ctime_check[i:i+batch_size]

                                # 순차 처리 (entry.stat()은 이미 캐싱됨, 병렬보다 오버헤드 적음)
                                batch_old_count = 0
                                for entry in batch:
                                    chart_num, file_date = check_entry_date(entry)
                                    if chart_num:
                                        chart_numbers.add(chart_num)
                                        ctime_matches += 1
                                        consecutive_old_files = 0
                                    elif file_date and file_date < self.today:
                                        batch_old_count += 1

                                # 이 배치에서 대부분 오래된 파일이면
                                if batch_old_count > len(batch) * 0.9:
                                    consecutive_old_files += 1

                                total_checked += len(batch)

                                # 진행 상황 로그
                                if total_checked % 2000 == 0 or i + batch_size >= len(need_ctime_check):
                                    log_callback(f"        ... {total_checked}/{len(need_ctime_check)} 확인 ({ctime_matches}건 발견)")

                                # 조기 종료: 연속 3배치가 모두 오래된 파일이면 중단
                                if consecutive_old_files >= 3:
                                    log_callback(f"     ⏹️  조기 종료: 최근 파일 없음 (총 {total_checked}개 확인)")
                                    break

                            log_callback(f"     ✅ 생성일 확인 완료: {ctime_matches}건 추가")

                            # 캐시 업데이트: 오늘 파일 제외한 모든 파일 저장
                            if HAS_CACHE:
                                # 오늘 생성된 파일을 제외한 나머지를 캐시에 추가
                                old_files = [e.name for e in candidate_entries if e.name not in chart_numbers]
                                update_cache_with_today_files(today_folder, old_files)
                                log_callback(f"     💾 캐시 업데이트 완료")

                        log_callback(f"     📊 최종 결과: {len(chart_numbers)}건 (중복 제외)")
                    return chart_numbers
                # scan_type == 'file'이 아닐 때는 아래 일반 스캔 로직으로 계속 진행

            # 오늘 폴더와 하위 폴더만 스캔 (os.walk 사용)
            log_callback(f"     📂 스캔 경로: {today_folder}")

            # 날짜 폴더가 없고 base_path를 스캔하는 경우 (실시간 파일/폴더)
            is_realtime_scan = (today_folder == base_path)

            total_files_count = 0
            total_dirs_count = 0


            # scan_type == 'both'이고 날짜 폴더 없을 때: 최상위 폴더+파일 스캔 (오늘 것만)
            if scan_type == 'both' and is_realtime_scan:
                log_callback(f"     🔍 최상위 폴더+파일 스캔 (정리 전, 오늘 mtime만)")

                valid_extensions = self.config['validation']['file_extensions']
                total_files_count = 0
                today_str = self.today.strftime('%Y%m%d')
                skipped_old = 0

                try:
                    items = os.listdir(today_folder)
                    for item in items:
                        item_path = os.path.join(today_folder, item)

                        # skip year/date folders
                        if os.path.isdir(item_path) and re.match(r'^(20\d{2}|\d{2}\.\d{2})$', item):
                            continue

                        # mtime filter: only today's items
                        try:
                            mtime = os.path.getmtime(item_path)
                            if time.strftime('%Y%m%d', time.localtime(mtime)) != today_str:
                                skipped_old += 1
                                continue
                        except OSError:
                            continue

                        if os.path.isdir(item_path):
                            total_dirs_count += 1
                            match = pattern.search(item)
                            if match:
                                chart_num = self.extract_chart_number(match)
                                if self.is_valid_chart_number(chart_num):
                                    chart_numbers.add(chart_num)

                        elif os.path.isfile(item_path):
                            if any(item.lower().endswith(ext) for ext in valid_extensions):
                                total_files_count += 1
                                match = pattern.search(item)
                                if match:
                                    chart_num = self.extract_chart_number(match)
                                    if self.is_valid_chart_number(chart_num):
                                        chart_numbers.add(chart_num)

                    log_callback(f"     📊 폴더: {total_dirs_count}개 / 파일: {total_files_count}개 / 매칭: {len(chart_numbers)}건 (과거 {skipped_old}개 스킵)")
                except Exception as e:
                    log_callback(f"     ❌ 스캔 오류: {e}")

            else:
                # 일반 스캔 (날짜 폴더가 있는 경우)
                for root, dirs, files in os.walk(today_folder):
                    total_files_count += len(files)
                    total_dirs_count += len(dirs)

                    # 파일 스캔
                    if scan_type in ['file', 'both']:
                        for file_name in files:
                            # 확장자 체크
                            if not any(file_name.lower().endswith(ext) for ext in self.config['validation']['file_extensions']):
                                continue

                            # 차트번호 추출
                            match = pattern.search(file_name)
                            if match:
                                chart_num = self.extract_chart_number(match)
                                if self.is_valid_chart_number(chart_num):
                                    chart_numbers.add(chart_num)

                    # 폴더 스캔 (OCT, HFA 등)
                    if scan_type == 'both':
                        for dir_name in dirs:
                            match = pattern.search(dir_name)
                            if match:
                                chart_num = self.extract_chart_number(match)
                                if self.is_valid_chart_number(chart_num):
                                    chart_numbers.add(chart_num)

            # 로그 출력 (실시간 스캔은 위에서 이미 출력)
            if not (scan_type == 'both' and is_realtime_scan):
                if scan_type == 'both':
                    log_callback(f"     📊 파일: {total_files_count}개 / 폴더: {total_dirs_count}개 / 매칭: {len(chart_numbers)}건")
                else:
                    log_callback(f"     📊 파일: {total_files_count}개 / 매칭: {len(chart_numbers)}건")

        except Exception as e:
            log_callback(f"  ❌ 오류: {equipment['name']} - {str(e)}")

        return chart_numbers

    def organize_directory(self, base_path, folder_structure, log_callback) -> int:
        """daehyuk.py 로직: 루트의 flat 항목을 날짜폴더로 정리 (배치 대체)"""
        if not os.path.exists(base_path):
            return 0

        target_folder = self._resolve_date_folder(folder_structure)
        target_path = os.path.join(base_path, target_folder)
        today_str = self.today.strftime('%Y%m%d')
        moved = 0

        for item in os.listdir(base_path):
            item_path = os.path.join(base_path, item)
            # skip year/month folders, log files, Thumbs.db
            if os.path.isdir(item_path) and re.match(r'^(20\d{2}|\d{2}\.\d{2})$', item):
                continue
            if item.startswith('Log') or item == 'Thumbs.db':
                continue
            # check modification date matches target date
            try:
                mtime = os.path.getmtime(item_path)
                if time.strftime('%Y%m%d', time.localtime(mtime)) != today_str:
                    continue
            except OSError:
                continue
            # create target folder and move
            if not os.path.exists(target_path):
                os.makedirs(target_path, exist_ok=True)
            try:
                dest = os.path.join(target_path, item)
                if os.path.exists(dest):
                    log_callback(f"    ⚠️ 이미 존재: {item}")
                    continue
                shutil.move(item_path, target_path)
                moved += 1
            except Exception as e:
                log_callback(f"    ❌ 이동 실패: {item} → {str(e)}")
                continue

        return moved

    def organize_all_directories(self, log_callback):
        """스캔 후 auto_organize 대상 디렉토리 자동 정리 (daehyuk.py 통합)"""
        targets = []

        for item in self.config.get('auto_organize', []):
            if item.get('path') and item.get('folder_structure'):
                targets.append((item['name'], item['path'], item['folder_structure']))

        total_moved = 0
        for name, path, fs in targets:
            moved = self.organize_directory(path, fs, log_callback)
            if moved > 0:
                log_callback(f"  📁 {name}: {moved}건 정리 완료")
                total_moved += moved

        return total_moved

    def count_sightmap(self, log_callback) -> int:
        """라식검사 카운트: Sightmap 고유 차트 수 = 라식 건수"""
        sm_config = self.config.get('sightmap', {})
        if not sm_config:
            return 0

        base_path = sm_config['path']
        folder_structure = sm_config.get('folder_structure', '')

        if not os.path.exists(base_path):
            log_callback(f"  ⚠️  Sightmap 경로 없음: {base_path}")
            return 0

        try:
            folder = self._resolve_date_folder(folder_structure)
            today_folder = os.path.join(base_path, folder)

            sm_pattern = re.compile(r'\s(\d+)-\d+')
            sightmap_charts = set()

            # 1) 날짜폴더 스캔
            if os.path.exists(today_folder):
                items = os.listdir(today_folder)
                for item in items:
                    match = sm_pattern.search(item)
                    if match:
                        sightmap_charts.add(match.group(1))
                log_callback(f"  📂 Sightmap (날짜폴더): {len(items)}건 (차트 {len(sightmap_charts)}명)")

            # 2) 루트에 아직 안 옮겨진 오늘 항목도 스캔
            today_str = self.today.strftime('%Y%m%d')
            root_count = 0
            for item in os.listdir(base_path):
                item_path = os.path.join(base_path, item)
                if item.startswith('20') and len(item) == 4 and os.path.isdir(item_path):
                    continue
                if item.startswith('Log') or item == 'Thumbs.db':
                    continue
                try:
                    mtime = os.path.getmtime(item_path)
                    if time.strftime('%Y%m%d', time.localtime(mtime)) != today_str:
                        continue
                except OSError:
                    continue
                root_count += 1
                match = sm_pattern.search(item)
                if match:
                    sightmap_charts.add(match.group(1))
            if root_count > 0:
                log_callback(f"  📂 Sightmap (루트): {root_count}건 추가")

            count = len(sightmap_charts)
            log_callback(f"  ✓ 라식검사: {count}건")

            return count

        except Exception as e:
            log_callback(f"  ❌ 라식검사 오류: {str(e)}")
            return 0

    def calculate_glaucoma(self, log_callback) -> int:
        """녹내장 계산 (HFA ∩ OCT)"""
        try:
            hfa_charts = self.chart_numbers.get('HFA', set())
            oct_charts = self.chart_numbers.get('OCT', set())
            glaucoma_charts = hfa_charts & oct_charts
            return len(glaucoma_charts)
        except Exception as e:
            log_callback(f"  ❌ 녹내장 계산 오류: {str(e)}")
            return 0

    def calculate_fundus(self, log_callback) -> int:
        """안저 계산 (Fundus + Secondary 폴더) - 최적화 버전"""
        fundus_charts = set()

        # 오늘 날짜 패턴
        today_str = self.today.strftime('%Y%m%d')
        today_str_dash = self.today.strftime('%Y-%m-%d')
        today_str_dot = self.today.strftime('%Y.%m.%d')
        date_patterns = [today_str, today_str_dash, today_str_dot]

        try:
            fundus_config = self.config['special_items']['안저']['folders']

            # 1. Fundus 폴더 처리 (날짜별 폴더 구조)
            if 'fundus' in fundus_config:
                fundus_info = fundus_config['fundus']
                base_path = fundus_info['path']
                pattern = re.compile(fundus_info['pattern'])

                log_callback(f"  📂 Fundus 스캔: {base_path}")

                if os.path.exists(base_path):
                    # 오늘 날짜 폴더 경로 생성
                    folder_structure = fundus_info.get('folder_structure', '')
                    today_folder = None

                    if folder_structure:
                        folder = self._resolve_date_folder(folder_structure)
                        today_folder = os.path.join(base_path, folder)

                    # 1) 날짜 폴더가 있으면 우선 스캔 (저녁 정리 후)
                    if today_folder and os.path.exists(today_folder):
                        log_callback(f"     📂 날짜 폴더: {today_folder}")
                        items = os.listdir(today_folder)
                        log_callback(f"     전체: {len(items)}개")

                        for item in items:
                            match = pattern.search(item)
                            if match:
                                chart_num = self.extract_chart_number(match)
                                if self.is_valid_chart_number(chart_num):
                                    fundus_charts.add(chart_num)

                        log_callback(f"     ✅ 날짜 폴더 매칭: {len(fundus_charts)}건")

                    # 2) 날짜 폴더가 없으면 base_path 스캔 (정리 전 파일)
                    # 매일 저녁 100% 정리하므로 최상위에 있는 것 = 오늘 것
                    if not today_folder or not os.path.exists(today_folder):
                        log_callback(f"     📂 최상위 경로 스캔: {base_path} (정리 전)")

                        try:
                            items = os.listdir(base_path)
                            # 하위 폴더 제외, 파일만
                            files = [f for f in items if os.path.isfile(os.path.join(base_path, f))]
                            log_callback(f"     전체 파일: {len(files)}개")

                            base_fundus_charts = set()
                            valid_extensions = self.config['validation']['file_extensions']

                            for file_name in files:
                                # 확장자 체크
                                if not any(file_name.lower().endswith(ext) for ext in valid_extensions):
                                    continue

                                # 패턴 매칭 (생성일 확인 없이)
                                match = pattern.search(file_name)
                                if match:
                                    chart_num = self.extract_chart_number(match)
                                    if self.is_valid_chart_number(chart_num):
                                        base_fundus_charts.add(chart_num)

                            if base_fundus_charts:
                                log_callback(f"     ✅ 최상위 파일 매칭: {len(base_fundus_charts)}건")
                                fundus_charts.update(base_fundus_charts)
                            else:
                                log_callback(f"     ⚠️  매칭된 파일 없음")
                        except Exception as e:
                            log_callback(f"     ❌ 최상위 경로 스캔 오류: {e}")
                else:
                    log_callback(f"  ⚠️  경로 없음: {base_path}")

            # 2. Secondary 폴더 처리 (파일명에 날짜 포함)
            if 'secondary' in fundus_config:
                secondary_info = fundus_config['secondary']
                folder_path = secondary_info['path']
                pattern = re.compile(secondary_info['pattern'])

                log_callback(f"  📂 Secondary 스캔: {folder_path}")

                if os.path.exists(folder_path):
                    try:
                        items = os.listdir(folder_path)
                        total_items = len(items)
                        log_callback(f"     전체: {total_items}개")

                        # 파일명에 오늘 날짜가 포함된 것만 필터링
                        # 예: 204775-20250919@161455-l4-s.jpg
                        filename_matched = 0
                        secondary_charts = set()

                        for item in items:
                            if today_str in item:  # 20251117 형식
                                filename_matched += 1
                                match = pattern.search(item)
                                if match:
                                    chart_num = self.extract_chart_number(match)
                                    if self.is_valid_chart_number(chart_num):
                                        secondary_charts.add(chart_num)

                        log_callback(f"     오늘 날짜 파일: {filename_matched}개")
                        log_callback(f"     ✅ Secondary: {len(secondary_charts)}명 (중복 제거)")

                        # 합집합
                        before_merge = len(fundus_charts)
                        fundus_charts.update(secondary_charts)
                        after_merge = len(fundus_charts)

                        if before_merge > 0:
                            overlap = before_merge + len(secondary_charts) - after_merge
                            if overlap > 0:
                                log_callback(f"     💡 Fundus & Secondary 중복: {overlap}명")

                    except Exception as e:
                        log_callback(f"  ⚠️  Secondary 스캔 오류: {e}")
                else:
                    log_callback(f"  ⚠️  경로 없음: {folder_path}")

        except Exception as e:
            log_callback(f"  ❌ 안저 계산 오류: {str(e)}")

        log_callback(f"  📊 안저 최종 집계: {len(fundus_charts)}명 (중복 제거 완료)")
        return len(fundus_charts)

    def _check_reservation_keywords(self, cell_value: str) -> Dict[str, bool]:
        """예약 셀에서 Verion/LensX/EX500 키워드 체크"""
        cell_lower = cell_value.lower()
        reservation = self.config['reservation']

        has_lensx = any(kw.lower() in cell_lower for kw in reservation['lensx_keywords'])

        if reservation.get('lensx_plus_pattern', False):
            if re.search(r'\+\s*[lL](?:ens)?[\s\.]*[xX]', cell_value):
                has_lensx = True

        has_verion = any(kw.lower() in cell_lower for kw in reservation['verion_keywords'])
        if has_lensx:
            has_verion = True

        has_ex500 = any(kw.lower() in cell_lower for kw in reservation['ex500_keywords'])

        return {'verion': has_verion, 'lensx': has_lensx, 'ex500': has_ex500}

    def process_reservation_file(self, file_path: str, log_callback) -> Dict[str, int]:
        """예약 파일 처리 (.xlsx, .xls 모두 지원)"""
        counts = {'verion': 0, 'lensx': 0, 'ex500': 0}
        found_cells = set()
        search_keyword = self.config['reservation'].get('search_keyword', '예약비고:').lower()

        try:
            # .xls 파일인 경우 xlrd로 읽기
            if file_path.lower().endswith('.xls') and not file_path.lower().endswith('.xlsx'):
                if not HAS_XLRD:
                    log_callback(f"  ⚠️  .xls 파일 읽기 실패: xlrd 라이브러리가 필요합니다")
                    log_callback(f"     설치: pip install xlrd")
                    return counts

                import xlrd
                xls_book = xlrd.open_workbook(file_path)

                for sheet in xls_book.sheets():
                    for row_idx in range(sheet.nrows):
                        for col_idx in range(sheet.ncols):
                            cell = sheet.cell(row_idx, col_idx)
                            if cell.value is None or cell.value == '':
                                continue

                            cell_value = str(cell.value)
                            if search_keyword not in cell_value.lower():
                                continue

                            cell_key = f"{sheet.name}_{row_idx}_{col_idx}_{cell_value}"
                            if cell_key in found_cells:
                                continue
                            found_cells.add(cell_key)

                            flags = self._check_reservation_keywords(cell_value)
                            for key in counts:
                                if flags[key]:
                                    counts[key] += 1

                return counts

            # .xlsx 파일은 openpyxl로 읽기
            wb = load_workbook(file_path, data_only=True)

            for sheet in wb.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is None:
                            continue

                        cell_value = str(cell.value)
                        if search_keyword not in cell_value.lower():
                            continue

                        cell_key = f"{sheet.title}_{cell.coordinate}_{cell_value}"
                        if cell_key in found_cells:
                            continue
                        found_cells.add(cell_key)

                        flags = self._check_reservation_keywords(cell_value)
                        for key in counts:
                            if flags[key]:
                                counts[key] += 1

            wb.close()

        except Exception as e:
            log_callback(f"  ❌ 예약 파일 처리 오류: {str(e)}")

        return counts

    def get_reservation_from_db(self, target_date: date, log_callback) -> Optional[Dict[str, int]]:
        """EMR DB(softcrm)에서 수술 예약 자동 조회 → Verion/LensX/EX500 카운트"""
        if not HAS_PYODBC:
            log_callback("  ⚠️  pyodbc 미설치 (DB 조회 불가)")
            return None

        db_config = self.config.get('emr_db', {})
        if not db_config.get('enabled', False):
            log_callback("  ⚠️  EMR DB 설정 없음 또는 비활성화")
            return None

        try:
            drivers = [d for d in pyodbc.drivers() if 'SQL Server' in d]
            log_callback(f"  📋 ODBC 드라이버: {drivers if drivers else '없음'}")
            if not drivers:
                log_callback("  ⚠️  SQL Server ODBC 드라이버 없음 (설치 필요)")
                return None
            driver = drivers[-1]

            log_callback(f"  🔗 DB 연결 시도: {db_config['server']} ({driver})")
            conn_str = (
                f"DRIVER={{{driver}}};"
                f"SERVER={db_config['server']};"
                f"UID={db_config['uid']};"
                f"PWD={db_config['pwd']};"
                f"DATABASE={db_config['database']};"
                f"TrustServerCertificate=yes"
            )
            conn = pyodbc.connect(conn_str, timeout=5)
            cursor = conn.cursor()

            date_str = target_date.strftime('%Y-%m-%d')
            cursor.execute("""
                SELECT OPERATIONR, OPERATIONL, COMMENT
                FROM RESERVATION
                WHERE RESERVE_DATE = ?
                  AND RESERVE_STATE != '2'
                  AND (ISNULL(OPERATIONR,'') != '' OR ISNULL(OPERATIONL,'') != '')
            """, date_str)

            reservation = self.config['reservation']
            counts = {'verion': 0, 'lensx': 0, 'ex500': 0}

            for row in cursor.fetchall():
                opr = (row.OPERATIONR or '').strip()
                opl = (row.OPERATIONL or '').strip()
                comment = (row.COMMENT or '').strip()
                combined = f"{opr} {opl} {comment}"

                flags = self._check_reservation_keywords(combined)
                for key in counts:
                    if flags[key]:
                        counts[key] += 1

            conn.close()
            return counts

        except Exception as e:
            log_callback(f"  ⚠️  DB 조회 오류: {str(e)}")
            return None

    def write_excel(self, output_path: str, staff_selected: List[str],
                   result_values: Dict[str, int], log_callback) -> bool:
        """엑셀 파일 작성 (result_values: 모든 항목의 확정된 값)"""
        try:
            template_file = get_template_path(self.config.get('template_file', ''))
            if not os.path.exists(template_file):
                log_callback(f"  ❌ 템플릿 파일 없음: {template_file}")
                log_callback(f"     (exe 옆 또는 번들에 '일일결산_템플릿.xlsx' 필요)")
                return False
            log_callback(f"  📄 템플릿: {template_file}")

            wb = load_workbook(template_file)
            ws = wb[self.config['target_sheet']]

            # 날짜 기입
            date_cell = self.config['date_cell']
            ws.cell(date_cell['row'], date_cell['col']).value = self.today.strftime('%Y-%m-%d')

            # 근무 인원 기입
            staff_cell = self.config['staff_cell']
            staff_count = len(staff_selected)
            staff_text = f"{staff_count}명( {', '.join(staff_selected)} )"
            ws.cell(staff_cell['row'], staff_cell['col']).value = staff_text

            # 각 장비별 결과 기입
            for equipment_id in self.config['equipment']:
                cell_info = self.config['equipment'][equipment_id]['cell']
                value = result_values.get(equipment_id, 0)
                ws.cell(cell_info['row'], cell_info['col']).value = value

            # 특수 항목 기입
            glaucoma_cell = self.config['special_items']['녹내장']['cell']
            ws.cell(glaucoma_cell['row'], glaucoma_cell['col']).value = result_values.get('GLAUCOMA', 0)

            fundus_cell = self.config['special_items']['안저']['cell']
            ws.cell(fundus_cell['row'], fundus_cell['col']).value = result_values.get('FUNDUS', 0)

            # Sightmap(라식) 자동 스캔 항목
            sm_config = self.config.get('sightmap', {})
            sightmap_cell = sm_config.get('cell', {'row': 10, 'col': 3})
            ws.cell(sightmap_cell['row'], sightmap_cell['col']).value = result_values.get('LASIK', 0)

            # 수기 입력 항목
            fag_cell = self.config['manual_input']['FAG']
            ws.cell(fag_cell['row'], fag_cell['col']).value = result_values.get('FAG', 0)

            glasses_cell = self.config['manual_input']['안경검사']
            ws.cell(glasses_cell['row'], glasses_cell['col']).value = result_values.get('GLASSES', 0)

            # 예약 파일 결과
            verion_cell = self.config['reservation']['cells']['verion']
            ws.cell(verion_cell['row'], verion_cell['col']).value = result_values.get('VERION', 0)

            lensx_cell = self.config['reservation']['cells']['lensx']
            ws.cell(lensx_cell['row'], lensx_cell['col']).value = result_values.get('LENSX', 0)

            ex500_cell = self.config['reservation']['cells']['ex500']
            ws.cell(ex500_cell['row'], ex500_cell['col']).value = result_values.get('EX500', 0)

            wb.save(output_path)
            wb.close()

            log_callback("  ✓ 엑셀 작성 완료")
            return True

        except Exception as e:
            log_callback(f"  ❌ 엑셀 작성 오류: {str(e)}")
            return False

    def convert_to_pdf(self, excel_path: str, pdf_path: str, log_callback) -> bool:
        """엑셀 파일을 PDF로 변환"""
        if not HAS_WIN32:
            log_callback("  ⚠️  pywin32가 없어 PDF 변환 불가")
            return False

        try:
            os.makedirs(os.path.dirname(pdf_path), exist_ok=True)

            # COM 라이브러리 초기화
            pythoncom.CoInitialize()

            try:
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False

                wb = excel.Workbooks.Open(os.path.abspath(excel_path))
                ws = wb.Worksheets(self.config['target_sheet'])

                ws.ExportAsFixedFormat(0, os.path.abspath(pdf_path))

                wb.Close(SaveChanges=False)
                excel.Quit()

                log_callback(f"  ✓ PDF 생성 완료: {pdf_path}")
                return True

            finally:
                # COM 라이브러리 정리
                pythoncom.CoUninitialize()

        except Exception as e:
            log_callback(f"  ❌ PDF 변환 오류: {str(e)}")
            return False


class DailyReportGUI:
    """일일결산 시스템의 GUI 클래스"""

    LABEL_MAP = {
        'OQAS': '백내장', 'HFA': '시야', 'OCT': 'OCT', 'ORB': 'ORB',
        'SP': '내피', 'TOPO': 'Tomey', 'GLAUCOMA': '녹내장', 'FUNDUS': '안저',
        'LASIK': '라식', 'GLASSES': '안경검사', 'FAG': 'FAG',
        'VERION': 'Verion', 'LENSX': 'LensX', 'EX500': 'EX500'
    }

    RESULT_ITEMS = [
        ('OQAS', '백내장'), ('HFA', '시야'), ('OCT', 'OCT'), ('ORB', 'ORB'),
        ('SP', '내피'), ('TOPO', 'Tomey'), ('GLAUCOMA', '녹내장'), ('FUNDUS', '안저'),
        ('LASIK', '라식'), ('GLASSES', '안경검사'), ('FAG', 'FAG'),
        ('VERION', 'Verion'), ('LENSX', 'LensX'), ('EX500', 'EX500'),
    ]

    def __init__(self, root: tk.Tk, system: DailyReportSystem):
        self.root = root
        self.system = system
        self.scan_results = {}
        self.setup_gui()

    def setup_gui(self):
        """GUI 구성 요소 생성"""
        self.root.title("일일결산 자동화 시스템 (최적화)")
        self.root.geometry("1400x800")
        self.root.resizable(True, True)

        # 메뉴바 추가
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        # 설정 메뉴
        settings_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="설정", menu=settings_menu)
        settings_menu.add_command(label="경로 설정...", command=self.open_settings)
        settings_menu.add_separator()
        settings_menu.add_command(label="설정 다시 로드", command=self.reload_config)

        # 도움말 메뉴
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="도움말", menu=help_menu)
        help_menu.add_command(label="정보", command=self.show_about)

        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

        # 3칸 레이아웃 설정
        main_frame.columnconfigure(0, weight=0)  # 왼쪽: 고정 너비
        main_frame.columnconfigure(1, weight=0)  # 중간: 고정 너비
        main_frame.columnconfigure(2, weight=1)  # 오른쪽: 가변 너비 (로그)
        main_frame.rowconfigure(0, weight=1)

        # 왼쪽 입력 영역 (날짜, 인원, 예약파일, 수기입력)
        left_frame = ttk.Frame(main_frame, padding="5")
        left_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # 중간 결과 영역 (스캔 결과)
        middle_frame = ttk.Frame(main_frame, padding="5")
        middle_frame.grid(row=0, column=1, sticky=(tk.W, tk.E, tk.N, tk.S))

        # 오른쪽 로그 영역
        right_frame = ttk.Frame(main_frame, padding="5")
        right_frame.grid(row=0, column=2, sticky=(tk.W, tk.E, tk.N, tk.S))
        right_frame.columnconfigure(0, weight=1)
        right_frame.rowconfigure(1, weight=1)

        # === 좌측 영역 구성 ===

        # 0. 결산 날짜 선택
        date_label = ttk.Label(left_frame, text="📅 결산 날짜", font=("", 12, "bold"))
        date_label.grid(row=0, column=0, columnspan=2, sticky=tk.W, pady=(0, 5))

        date_frame = ttk.Frame(left_frame)
        date_frame.grid(row=1, column=0, columnspan=2, sticky=tk.W, pady=(0, 10))

        # 날짜 입력 (YYYY-MM-DD)
        self.date_entry = ttk.Entry(date_frame, width=12)
        self.date_entry.insert(0, date.today().strftime('%Y-%m-%d'))
        self.date_entry.grid(row=0, column=0, padx=(0, 5))

        today_btn = ttk.Button(date_frame, text="오늘", width=6,
                               command=lambda: self.set_date(date.today()))
        today_btn.grid(row=0, column=1, padx=2)

        yesterday_btn = ttk.Button(date_frame, text="어제", width=6,
                                   command=lambda: self.set_date(date.today() - timedelta(days=1)))
        yesterday_btn.grid(row=0, column=2, padx=2)

        ttk.Separator(left_frame, orient='horizontal').grid(row=2, column=0, columnspan=2,
                                                             sticky=(tk.W, tk.E), pady=5)

        # 1. 근무 인원 (자동 로드)
        staff_header_frame = ttk.Frame(left_frame)
        staff_header_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 5))

        staff_label = ttk.Label(staff_header_frame, text="👥 근무 인원", font=("", 12, "bold"))
        staff_label.grid(row=0, column=0, sticky=tk.W)

        # 새로고침 버튼 (Enter 키로도 가능)
        refresh_btn = ttk.Button(staff_header_frame, text="🔄", width=3,
                                command=self.refresh_staff_list)
        refresh_btn.grid(row=0, column=1, padx=5)

        # API 상태 라벨
        self.api_status_label = ttk.Label(staff_header_frame, text="", foreground="gray", font=("", 9))
        self.api_status_label.grid(row=0, column=2, sticky=tk.W)

        # 인원 수 라벨
        self.staff_count_label = ttk.Label(staff_header_frame, text="", foreground="blue", font=("", 10, "bold"))
        self.staff_count_label.grid(row=0, column=3, padx=(10, 0), sticky=tk.W)

        # 직원 체크박스 프레임 (스크롤 가능)
        staff_canvas = tk.Canvas(left_frame, height=100)
        staff_scrollbar = ttk.Scrollbar(left_frame, orient="vertical", command=staff_canvas.yview)
        self.staff_scrollable = ttk.Frame(staff_canvas)

        self.staff_scrollable.bind(
            "<Configure>",
            lambda e: staff_canvas.configure(scrollregion=staff_canvas.bbox("all"))
        )

        staff_canvas.create_window((0, 0), window=self.staff_scrollable, anchor="nw")
        staff_canvas.configure(yscrollcommand=staff_scrollbar.set)

        staff_canvas.grid(row=4, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        staff_scrollbar.grid(row=4, column=1, sticky=(tk.N, tk.S), pady=(0, 10))

        # 직원 체크박스 변수들
        self.staff_vars = {}
        self.current_staff = []
        self.refresh_staff_list()

        # 날짜 입력 필드에 Enter 키 바인딩
        self.date_entry.bind('<Return>', lambda e: self.refresh_staff_list())

        # 2. 수기 입력 (FAG, 안경검사, OCTS만 남음)
        ttk.Separator(left_frame, orient='horizontal').grid(row=5, column=0, columnspan=2,
                                                             sticky=(tk.W, tk.E), pady=10)

        manual_label = ttk.Label(left_frame, text="✍ 수기 입력", font=("", 12, "bold"))
        manual_label.grid(row=6, column=0, columnspan=2, sticky=tk.W, pady=(0, 10))

        fag_label = ttk.Label(left_frame, text="FAG:")
        fag_label.grid(row=7, column=0, sticky=tk.W, padx=(0, 5))

        self.fag_entry = ttk.Entry(left_frame, width=10)
        self.fag_entry.insert(0, "0")
        self.fag_entry.grid(row=7, column=1, sticky=tk.W, pady=3)

        glasses_label = ttk.Label(left_frame, text="안경검사:")
        glasses_label.grid(row=8, column=0, sticky=tk.W, padx=(0, 5))

        self.glasses_entry = ttk.Entry(left_frame, width=10)
        self.glasses_entry.insert(0, "0")
        self.glasses_entry.grid(row=8, column=1, sticky=tk.W, pady=3)

        octs_label = ttk.Label(left_frame, text="OCTS:")
        octs_label.grid(row=9, column=0, sticky=tk.W, padx=(0, 5))

        self.octs_entry = ttk.Entry(left_frame, width=10)
        self.octs_entry.insert(0, "0")
        self.octs_entry.grid(row=9, column=1, sticky=tk.W, pady=3)

        # 3. 스캔 버튼
        ttk.Separator(left_frame, orient='horizontal').grid(row=10, column=0, columnspan=2,
                                                             sticky=(tk.W, tk.E), pady=10)

        self.scan_button = ttk.Button(left_frame, text="🔍 스캔 시작", command=self.run_scan)
        self.scan_button.grid(row=11, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=8)

        # === 중간 영역 구성 (스캔 결과) ===

        result_label = ttk.Label(middle_frame, text="📊 스캔 결과 (수정 가능)", font=("", 12, "bold"))
        result_label.grid(row=0, column=0, columnspan=2, sticky=tk.W, pady=(0, 10))

        # 결과 프레임 (스크롤 가능)
        result_canvas = tk.Canvas(middle_frame, height=600)
        result_scrollbar = ttk.Scrollbar(middle_frame, orient="vertical", command=result_canvas.yview)
        self.result_frame = ttk.Frame(result_canvas)

        self.result_frame.bind(
            "<Configure>",
            lambda e: result_canvas.configure(scrollregion=result_canvas.bbox("all"))
        )

        result_canvas.create_window((0, 0), window=self.result_frame, anchor="nw")
        result_canvas.configure(yscrollcommand=result_scrollbar.set)

        result_canvas.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        result_scrollbar.grid(row=1, column=1, sticky=(tk.N, tk.S), pady=(0, 10))

        # 결과 항목들 (Entry 위젯) - 초기에는 비활성화
        self.result_entries = {}

        for idx, (key, label_text) in enumerate(self.RESULT_ITEMS):
            label = ttk.Label(self.result_frame, text=f"{label_text}:")
            label.grid(row=idx, column=0, sticky=tk.W, padx=(0, 10), pady=3)

            entry = ttk.Entry(self.result_frame, width=12, state='disabled')
            entry.insert(0, "0")
            entry.grid(row=idx, column=1, sticky=tk.W, pady=3)
            self.result_entries[key] = entry

        # PDF 출력 버튼 (초기에는 비활성화)
        self.output_button = ttk.Button(middle_frame, text="✅ 확정 및 PDF 출력",
                                        command=self.run_output, state='disabled')
        self.output_button.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=10)

        # === 오른쪽 영역 구성 (실행 로그) ===

        log_label = ttk.Label(right_frame, text="📋 실행 로그", font=("", 12, "bold"))
        log_label.grid(row=0, column=0, sticky=tk.W, pady=(0, 10))

        self.log_text = scrolledtext.ScrolledText(right_frame, width=70, height=42,
                                                   state='disabled', wrap=tk.WORD)
        self.log_text.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

    def set_date(self, target_date: date):
        """날짜 설정"""
        self.date_entry.delete(0, tk.END)
        self.date_entry.insert(0, target_date.strftime('%Y-%m-%d'))
        # 날짜 변경 시 직원 목록도 새로고침
        self.refresh_staff_list()

    def load_staff_list(self, all_staff: List[str] = None, working_staff: List[str] = None):
        """직원 체크박스 생성 (근무 인원은 체크, 휴무 인원은 체크 해제)"""
        # 기존 체크박스 제거
        for widget in self.staff_scrollable.winfo_children():
            widget.destroy()

        # fallback: config에서 가져오기 (전원 체크)
        if all_staff is None:
            all_staff = self.system.config.get('staff_list', [])
        if working_staff is None:
            working_staff = all_staff

        # 현재 근무 인원 저장
        self.current_staff = working_staff

        # 인원 수 표시
        self.staff_count_label.config(text=f"{len(working_staff)}명")

        # 체크박스 생성 (2열로 배치)
        self.staff_vars = {}
        for i, staff_name in enumerate(all_staff):
            var = tk.BooleanVar(value=(staff_name in working_staff))
            self.staff_vars[staff_name] = var
            cb = ttk.Checkbutton(self.staff_scrollable, text=staff_name, variable=var,
                                command=self.update_staff_count)
            cb.grid(row=i // 2, column=i % 2, sticky=tk.W, padx=5, pady=1)

    def update_staff_count(self):
        """체크된 직원 수 업데이트"""
        count = sum(1 for var in self.staff_vars.values() if var.get())
        self.staff_count_label.config(text=f"{count}명")

    def refresh_staff_list(self):
        """Schedule API에서 직원 목록 자동 로드 (휴무자 제외)"""
        try:
            # 날짜 파싱
            date_str = self.date_entry.get()
            try:
                target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                self.api_status_label.config(text="날짜 형식 오류", foreground="red")
                self.load_staff_list()
                return

            # API에서 직원 목록 가져오기
            result = self.system.get_staff_from_api(target_date, "검사실")

            if result:
                all_staff, working_staff = result
                off_count = len(all_staff) - len(working_staff)
                self.load_staff_list(all_staff, working_staff)
                status = f"✓ API ({len(working_staff)}/{len(all_staff)}명"
                if off_count > 0:
                    status += f", 휴무 {off_count}명"
                status += ")"
                self.api_status_label.config(text=status, foreground="green")
            else:
                # API 실패 → config 전체 사용 (실제 에러 표시)
                self.load_staff_list()
                error_detail = getattr(self.system, '_last_api_error', None) or '알 수 없는 오류'
                self.api_status_label.config(
                    text=f"⚠ API 실패: {error_detail}",
                    foreground="orange"
                )
        except Exception as e:
            # 예외 발생 → config 전체 사용
            self.load_staff_list()
            self.api_status_label.config(text=f"⚠ 오류: {str(e)[:50]}", foreground="red")

    def log(self, message: str):
        """로그 메시지 출력 (화면 + 파일) - 스레드 안전"""
        def _update_ui():
            self.log_text.configure(state='normal')
            self.log_text.insert(tk.END, message + '\n')
            self.log_text.see(tk.END)
            self.log_text.configure(state='disabled')

        # 백그라운드 스레드에서도 안전하게 UI 업데이트
        self.root.after(0, _update_ui)

    def get_selected_staff(self) -> List[str]:
        """체크된 직원 목록 반환"""
        return [name for name, var in self.staff_vars.items() if var.get()]

    def run_scan(self):
        """1단계: 스캔 실행"""
        self.scan_button.config(state='disabled')

        self.log_text.configure(state='normal')
        self.log_text.delete(1.0, tk.END)
        self.log_text.configure(state='disabled')

        thread = threading.Thread(target=self.process_scan, daemon=True)
        thread.start()

    def run_output(self):
        """2단계: PDF 출력"""
        self.output_button.config(state='disabled')

        thread = threading.Thread(target=self.process_output, daemon=True)
        thread.start()

    def process_scan(self):
        """1단계: 스캔 처리 - 결과를 화면에 표시"""

        try:
            # 날짜 파싱
            date_str = self.date_entry.get()
            try:
                target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                self.system.today = target_date
            except ValueError:
                self.log("❌ 날짜 형식 오류! YYYY-MM-DD 형식으로 입력하세요.")
                self.scan_button.config(state='normal')
                return

            self.log("=" * 54)
            self.log(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 스캔 시작")
            self.log(f"결산 날짜: {target_date.strftime('%Y-%m-%d')}")
            self.log("=" * 54)
            self.log("")

            # 1. 파일 자동 분류 (루트 → 날짜폴더 정리, 스캔 전에 실행)
            self.log("[1/5] 파일 자동 분류 중...")
            organized = self.system.organize_all_directories(self.log)
            if organized > 0:
                self.log(f"  ✓ 총 {organized}건 날짜폴더로 정리 완료")
            else:
                self.log(f"  ✓ 정리할 항목 없음 (이미 분류됨)")

            self.log("")

            # 2. 디렉토리 자동 스캔 (날짜폴더에서 카운트)
            self.log("[2/5] 디렉토리 자동 스캔 중...")
            for equipment_id in self.system.config['equipment'].keys():
                equipment_name = self.system.config['equipment'][equipment_id]['name']
                self.log(f"  🔍 {equipment_name} 스캔 중...")

                chart_set = self.system.scan_directory_fast(equipment_id, self.log)
                self.system.chart_numbers[equipment_id] = chart_set

                self.log(f"  ✓ {equipment_name}: {len(chart_set)}건")

            self.log("")

            # 3. 특수 항목 계산
            self.log("[3/5] 특수 항목 계산 중...")

            glaucoma_count = self.system.calculate_glaucoma(self.log)
            self.log(f"  ✓ 녹내장 (HFA ∩ OCT): {glaucoma_count}건")

            fundus_count = self.system.calculate_fundus(self.log)
            self.log(f"  ✓ 안저: {fundus_count}건")

            self.log("")

            # 4. Sightmap(라식) 자동 카운트
            self.log("[4/5] Sightmap(라식) 스캔 중...")
            sightmap_count = self.system.count_sightmap(self.log)

            self.log("")

            # 5. 수술 예약 조회 (DB 자동)
            reservation_counts = {'verion': 0, 'lensx': 0, 'ex500': 0}

            self.log("[5/5] 수술 예약 조회 중...")
            db_counts = self.system.get_reservation_from_db(target_date, self.log)

            if db_counts is not None:
                reservation_counts = db_counts
                self.log(f"  ✓ EMR DB 자동 조회 완료")
                self.log(f"  ✓ Verion: {reservation_counts['verion']}건")
                self.log(f"  ✓ LensX: {reservation_counts['lensx']}건")
                self.log(f"  ✓ EX500: {reservation_counts['ex500']}건")
            else:
                self.log("  ⚠️  DB 연결 실패 (결과 화면에서 수동 입력 가능)")

            self.log("")

            # 스캔 결과를 인스턴스 변수에 저장
            self.scan_results = {
                'glaucoma_count': glaucoma_count,
                'fundus_count': fundus_count,
                'sightmap_count': sightmap_count,
                'reservation_counts': reservation_counts
            }

            # 결과 Entry 위젯 업데이트
            self.root.after(0, self.update_result_entries)

            self.log("")
            self.log("=" * 54)
            self.log("✅ 스캔 완료! 결과를 확인하고 수정 후 PDF 출력 버튼을 클릭하세요.")
            self.log("=" * 54)
            self.log("")

        except Exception as e:
            self.log("")
            self.log("=" * 54)
            self.log(f"❌ 오류 발생: {str(e)}")
            self.log("=" * 54)
            self.scan_button.config(state='normal')
    
    def update_result_entries(self):
        """스캔 결과를 Entry 위젯에 표시하고 편집 가능하게 설정"""
        # 각 항목의 값 설정
        entry_values = {
            'OQAS': len(self.system.chart_numbers.get('OQAS', set())),
            'HFA': len(self.system.chart_numbers.get('HFA', set())),
            'OCT': len(self.system.chart_numbers.get('OCT', set())) + int(self.octs_entry.get() or 0),
            'ORB': len(self.system.chart_numbers.get('ORB', set())),
            'SP': len(self.system.chart_numbers.get('SP', set())),
            'TOPO': len(self.system.chart_numbers.get('TOPO', set())),
            'GLAUCOMA': self.scan_results['glaucoma_count'],
            'FUNDUS': self.scan_results['fundus_count'],
            'LASIK': self.scan_results['sightmap_count'],
            'GLASSES': int(self.glasses_entry.get() or 0),
            'FAG': int(self.fag_entry.get() or 0),
            'VERION': self.scan_results['reservation_counts']['verion'],
            'LENSX': self.scan_results['reservation_counts']['lensx'],
            'EX500': self.scan_results['reservation_counts']['ex500'],
        }

        # Entry 위젯 업데이트 및 편집 가능하게 설정
        for key, value in entry_values.items():
            entry = self.result_entries[key]
            entry.config(state='normal')
            entry.delete(0, tk.END)
            entry.insert(0, str(value))

        # PDF 출력 버튼 활성화
        self.output_button.config(state='normal')
        self.scan_button.config(state='normal')

    def process_output(self):
        """2단계: PDF 출력 - Entry 위젯의 값을 읽어서 엑셀/PDF 생성"""

        try:
            self.log("")
            self.log("=" * 54)
            self.log(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] PDF 출력 시작")
            self.log("=" * 54)
            self.log("")

            # Entry 위젯에서 값 읽기
            self.log("[1/2] 확정된 값:")
            try:
                result_values = {}
                for key, entry in self.result_entries.items():
                    value = int(entry.get() or 0)
                    result_values[key] = value
                    self.log(f"  {self.LABEL_MAP.get(key, key)}: {value}건")
            except ValueError as e:
                self.log(f"  ⚠️  값 읽기 오류: {e}")
                self.output_button.config(state='normal')
                return

            self.log("")

            # 엑셀 작성용 데이터 준비
            staff_selected = self.get_selected_staff()
            if not staff_selected:
                self.log("  ⚠️  경고: 직원이 선택되지 않았습니다.")

            # 엑셀 작성
            self.log("[2/2] 엑셀 파일 작성 및 PDF 생성 중...")

            target_date = self.system.today
            today_str = target_date.strftime('%Y%m%d')
            temp_excel = f"일일결산_{today_str}_temp.xlsx"

            success = self.system.write_excel(
                temp_excel, staff_selected, result_values, self.log
            )

            if not success:
                self.log("")
                self.log("=" * 54)
                self.log("❌ 결산 실패: 엑셀 작성 오류")
                self.log("=" * 54)
                self.output_button.config(state='normal')
                return

            self.log("")

            # PDF 변환 - 출력 경로 자동 결정
            pdf_output_config = self.system.config.get('output_pdf', '')
            if pdf_output_config:
                pdf_resolved = pdf_output_config.replace('{date}', today_str)
                if os.path.isabs(pdf_resolved):
                    # 절대경로: 드라이브 접근 가능한지 확인
                    drive = os.path.splitdrive(pdf_resolved)[0]
                    if drive and os.path.exists(drive + os.sep):
                        pdf_path = pdf_resolved
                    else:
                        pdf_path = os.path.join(get_exe_dir(), 'PDF', f'일일결산_{today_str}.pdf')
                else:
                    # 상대경로: exe 위치 기준으로 resolve
                    pdf_path = os.path.join(get_exe_dir(), pdf_resolved)
            else:
                pdf_path = os.path.join(get_exe_dir(), 'PDF', f'일일결산_{today_str}.pdf')

            # PDF 출력 디렉토리 생성
            os.makedirs(os.path.dirname(pdf_path), exist_ok=True)

            pdf_success = self.system.convert_to_pdf(temp_excel, pdf_path, self.log)

            # 결산 완료 후 PDF를 일일결산\YYYY\MM\MM.DD.pdf로 정리
            if pdf_success and os.path.exists(pdf_path):
                archive_path = self._organize_settlement_pdf(pdf_path, target_date)

            self.log("")
            self.log("=" * 54)
            self.log("✅ 결산 완료!")
            self.log("=" * 54)
            self.log("")

            # PDF 열기
            if pdf_success and os.path.exists(pdf_path):
                self.log("📄 PDF 파일을 엽니다...")
                if sys.platform == 'win32':
                    os.startfile(pdf_path)
                else:
                    self.log(f"  PDF 경로: {pdf_path}")

                try:
                    os.remove(temp_excel)
                except:
                    pass
            else:
                self.log(f"📄 엑셀 파일이 저장되었습니다: {temp_excel}")

        except Exception as e:
            self.log("")
            self.log("=" * 54)
            self.log(f"❌ 오류 발생: {str(e)}")
            self.log("=" * 54)

        finally:
            self.output_button.config(state='normal')

    def _organize_settlement_pdf(self, pdf_path: str, target_date: date) -> Optional[str]:
        """결산 PDF를 일일결산\YYYY\MM\MM.DD.pdf로 정리 (검사폴더 정리와 동일 패턴)"""
        try:
            # PDF 폴더의 상위 = 결산 기본 디렉토리
            pdf_dir = os.path.dirname(pdf_path)
            base_dir = os.path.dirname(pdf_dir)  # PDF 폴더의 상위

            archive_dir = os.path.join(base_dir, '일일결산',
                                       target_date.strftime('%Y'),
                                       target_date.strftime('%m'))
            os.makedirs(archive_dir, exist_ok=True)

            archive_name = target_date.strftime('%m.%d') + '.pdf'
            archive_path = os.path.join(archive_dir, archive_name)

            shutil.copy2(pdf_path, archive_path)
            self.log(f"  📁 결산 정리: 일일결산/{target_date.strftime('%Y/%m/%m.%d')}.pdf")
            return archive_path

        except Exception as e:
            self.log(f"  ⚠️  결산 정리 실패: {str(e)}")
            return None

    def open_settings(self):
        """설정 창 열기"""
        settings_window = tk.Toplevel(self.root)
        settings_window.title("경로 설정")
        settings_window.geometry("800x600")
        settings_window.resizable(True, True)

        # 스크롤 가능한 프레임 생성
        canvas = tk.Canvas(settings_window)
        scrollbar = ttk.Scrollbar(settings_window, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # 설정 항목 저장용 딕셔너리
        path_entries = {}

        row = 0

        # 1. 템플릿 파일 경로
        ttk.Label(scrollable_frame, text="📄 템플릿 파일", font=("", 11, "bold")).grid(
            row=row, column=0, columnspan=3, sticky=tk.W, pady=(10, 5), padx=10
        )
        row += 1

        ttk.Label(scrollable_frame, text="템플릿 경로:").grid(row=row, column=0, sticky=tk.W, padx=(20, 5))
        template_entry = ttk.Entry(scrollable_frame, width=50)
        template_entry.insert(0, self.system.config.get('template_file', ''))
        template_entry.grid(row=row, column=1, sticky=(tk.W, tk.E), padx=5)

        def browse_template():
            filename = filedialog.askopenfilename(
                title="템플릿 파일 선택",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if filename:
                template_entry.delete(0, tk.END)
                template_entry.insert(0, filename)

        ttk.Button(scrollable_frame, text="찾아보기", command=browse_template, width=10).grid(
            row=row, column=2, padx=5
        )
        path_entries['template_file'] = template_entry
        row += 1

        ttk.Separator(scrollable_frame, orient='horizontal').grid(
            row=row, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10, padx=10
        )
        row += 1

        # 2. 출력 PDF 경로
        ttk.Label(scrollable_frame, text="📤 출력 PDF", font=("", 11, "bold")).grid(
            row=row, column=0, columnspan=3, sticky=tk.W, pady=(10, 5), padx=10
        )
        row += 1

        ttk.Label(scrollable_frame, text="PDF 경로:").grid(row=row, column=0, sticky=tk.W, padx=(20, 5))
        pdf_entry = ttk.Entry(scrollable_frame, width=50)
        pdf_entry.insert(0, self.system.config.get('output_pdf', ''))
        pdf_entry.grid(row=row, column=1, sticky=(tk.W, tk.E), padx=5)

        def browse_pdf():
            dirname = filedialog.askdirectory(title="PDF 저장 폴더 선택")
            if dirname:
                current = pdf_entry.get()
                filename = os.path.basename(current) if current else "일일결산_{date}.pdf"
                new_path = os.path.join(dirname, filename)
                pdf_entry.delete(0, tk.END)
                pdf_entry.insert(0, new_path)

        ttk.Button(scrollable_frame, text="찾아보기", command=browse_pdf, width=10).grid(
            row=row, column=2, padx=5
        )
        path_entries['output_pdf'] = pdf_entry
        row += 1

        ttk.Separator(scrollable_frame, orient='horizontal').grid(
            row=row, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10, padx=10
        )
        row += 1

        # 3. 장비 경로들
        ttk.Label(scrollable_frame, text="🔬 장비 경로", font=("", 11, "bold")).grid(
            row=row, column=0, columnspan=3, sticky=tk.W, pady=(10, 5), padx=10
        )
        row += 1

        for eq_id, eq_info in self.system.config['equipment'].items():
            eq_name = eq_info['name']
            ttk.Label(scrollable_frame, text=f"{eq_name} ({eq_id}):").grid(
                row=row, column=0, sticky=tk.W, padx=(20, 5)
            )

            entry = ttk.Entry(scrollable_frame, width=50)
            entry.insert(0, eq_info.get('path', ''))
            entry.grid(row=row, column=1, sticky=(tk.W, tk.E), padx=5)

            def make_browse(eq_id_local, entry_local):
                def browse():
                    dirname = filedialog.askdirectory(title=f"{eq_id_local} 경로 선택")
                    if dirname:
                        entry_local.delete(0, tk.END)
                        entry_local.insert(0, dirname)
                return browse

            ttk.Button(
                scrollable_frame, text="찾아보기",
                command=make_browse(eq_id, entry), width=10
            ).grid(row=row, column=2, padx=5)

            path_entries[f'equipment.{eq_id}.path'] = entry
            row += 1

        ttk.Separator(scrollable_frame, orient='horizontal').grid(
            row=row, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10, padx=10
        )
        row += 1

        # 4. 안저 폴더 경로
        ttk.Label(scrollable_frame, text="👁 안저 경로", font=("", 11, "bold")).grid(
            row=row, column=0, columnspan=3, sticky=tk.W, pady=(10, 5), padx=10
        )
        row += 1

        fundus_folders = self.system.config['special_items']['안저'].get('folders', {})
        for folder_id, folder_info in fundus_folders.items():
            ttk.Label(scrollable_frame, text=f"{folder_id}:").grid(
                row=row, column=0, sticky=tk.W, padx=(20, 5)
            )

            entry = ttk.Entry(scrollable_frame, width=50)
            entry.insert(0, folder_info.get('path', ''))
            entry.grid(row=row, column=1, sticky=(tk.W, tk.E), padx=5)

            def make_browse_fundus(folder_id_local, entry_local):
                def browse():
                    dirname = filedialog.askdirectory(title=f"안저 {folder_id_local} 경로 선택")
                    if dirname:
                        entry_local.delete(0, tk.END)
                        entry_local.insert(0, dirname)
                return browse

            ttk.Button(
                scrollable_frame, text="찾아보기",
                command=make_browse_fundus(folder_id, entry), width=10
            ).grid(row=row, column=2, padx=5)

            path_entries[f'fundus.{folder_id}.path'] = entry
            row += 1

        # 버튼 프레임
        button_frame = ttk.Frame(settings_window)
        button_frame.pack(side="bottom", fill="x", padx=10, pady=10)

        def save_settings():
            """설정 저장"""
            try:
                # 템플릿 및 PDF 경로
                self.system.config['template_file'] = path_entries['template_file'].get()
                self.system.config['output_pdf'] = path_entries['output_pdf'].get()

                # 장비 경로
                for key, entry in path_entries.items():
                    if key.startswith('equipment.'):
                        _, eq_id, _ = key.split('.')
                        self.system.config['equipment'][eq_id]['path'] = entry.get()
                    elif key.startswith('fundus.'):
                        _, folder_id, _ = key.split('.')
                        self.system.config['special_items']['안저']['folders'][folder_id]['path'] = entry.get()

                # config.json 저장 (exe 옆에 외부 파일로)
                if getattr(sys, 'frozen', False):
                    save_path = os.path.join(os.path.dirname(sys.executable), 'config.json')
                else:
                    save_path = self.system.config_path
                with open(save_path, 'w', encoding='utf-8') as f:
                    json.dump(self.system.config, f, indent=2, ensure_ascii=False)

                messagebox.showinfo("성공", "설정이 저장되었습니다.")
                settings_window.destroy()

            except Exception as e:
                messagebox.showerror("오류", f"설정 저장 중 오류 발생:\n{str(e)}")

        ttk.Button(button_frame, text="저장", command=save_settings, width=15).pack(side="right", padx=5)
        ttk.Button(button_frame, text="취소", command=settings_window.destroy, width=15).pack(side="right")

        # 스크롤 가능하게 마우스 휠 이벤트 바인딩
        def on_mousewheel(event):
            try:
                if canvas.winfo_exists():
                    canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            except:
                pass

        canvas.bind_all("<MouseWheel>", on_mousewheel)

        # 창 닫힐 때 이벤트 바인딩 해제
        def on_close():
            canvas.unbind_all("<MouseWheel>")
            settings_window.destroy()

        settings_window.protocol("WM_DELETE_WINDOW", on_close)

    def reload_config(self):
        """설정 다시 로드"""
        try:
            config_path = get_config_path()
            with open(config_path, 'r', encoding='utf-8') as f:
                self.system.config = json.load(f)
            self.system.config_path = config_path
            messagebox.showinfo("성공", "설정을 다시 로드했습니다.")
        except Exception as e:
            messagebox.showerror("오류", f"설정 로드 중 오류 발생:\n{str(e)}")

    def show_about(self):
        """정보 표시"""
        about_text = """일일결산 자동화 시스템 v2.0

안과 검사실의 일일 통계를 자동으로 수집하고
PDF 보고서를 생성하는 프로그램입니다.

주요 기능:
• 장비별 자동 스캔 (SP, TOPO, ORB, OCT, HFA, OQAS)
• 특수 항목 계산 (녹내장, 안저)
• 예약 파일 분석 (Verion, LensX, EX500)
• 2단계 워크플로우 (스캔 → 확인/수정 → PDF 출력)
• 경로 설정 기능

© 2025 일일결산 자동화 시스템
"""
        messagebox.showinfo("일일결산 자동화 시스템", about_text)


def main():
    """메인 함수"""
    config_path = get_config_path()
    if not os.path.exists(config_path):
        messagebox.showerror("오류", "config.json 파일을 찾을 수 없습니다.\nexe 옆에 config.json을 두거나 번들에 포함되어 있어야 합니다.")
        sys.exit(1)

    system = DailyReportSystem(config_path)

    root = tk.Tk()
    app = DailyReportGUI(root, system)
    root.mainloop()


if __name__ == "__main__":
    main()
