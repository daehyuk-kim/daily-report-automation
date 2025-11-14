#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
일일결산 자동화 시스템
안과 검사실의 일일 통계를 자동으로 수집하고 PDF 보고서를 생성하는 프로그램
"""

import os
import sys
import json
import re
import threading
from pathlib import Path
from datetime import datetime, date
from typing import Set, Dict, List, Optional, Tuple
from concurrent.futures import ThreadPoolExecutor, as_completed
import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext, messagebox

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("오류: openpyxl이 설치되지 않았습니다.")
    print("설치: pip install openpyxl")
    sys.exit(1)

try:
    import pandas as pd
except ImportError:
    print("오류: pandas가 설치되지 않았습니다.")
    print("설치: pip install pandas")
    sys.exit(1)

# Windows에서만 pywin32 임포트
if sys.platform == 'win32':
    try:
        import win32com.client
        HAS_WIN32 = True
    except ImportError:
        HAS_WIN32 = False
        print("경고: pywin32가 설치되지 않았습니다. PDF 변환을 사용할 수 없습니다.")
        print("설치: pip install pywin32")
else:
    HAS_WIN32 = False


class DailyReportSystem:
    """일일결산 시스템의 메인 클래스"""

    def __init__(self, config_path: str = "config.json"):
        """
        시스템 초기화

        Args:
            config_path: 설정 파일 경로
        """
        self.config = self.load_config(config_path)
        self.chart_numbers = {}  # 각 장비별 차트번호 저장
        self.results = {}  # 최종 결과 저장

    def load_config(self, config_path: str) -> dict:
        """설정 파일 로드"""
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except FileNotFoundError:
            messagebox.showerror("오류", f"설정 파일을 찾을 수 없습니다: {config_path}")
            sys.exit(1)
        except json.JSONDecodeError:
            messagebox.showerror("오류", "설정 파일 형식이 올바르지 않습니다.")
            sys.exit(1)

    def is_valid_chart_number(self, chart_num_str: str) -> bool:
        """
        차트번호 유효성 검증

        Args:
            chart_num_str: 차트번호 문자열

        Returns:
            유효 여부
        """
        try:
            # 선행 0 체크
            if chart_num_str.startswith('0') and len(chart_num_str) > 1:
                return False

            chart_num = int(chart_num_str)
            min_val = self.config['validation']['chart_number_min']
            max_val = self.config['validation']['chart_number_max']

            return min_val <= chart_num <= max_val
        except (ValueError, KeyError):
            return False

    def is_today_file(self, file_path: Path, date_filter: Optional[str] = None) -> bool:
        """
        파일이 오늘 생성되었는지 확인

        Args:
            file_path: 파일 경로
            date_filter: 날짜 필터 (YYYY-MM-DD 형식)

        Returns:
            오늘 생성 여부
        """
        try:
            # 파일 생성 시간
            ctime = os.path.getctime(file_path)
            file_date = date.fromtimestamp(ctime)

            # 오늘 날짜 확인
            today = date.today()
            if file_date != today:
                return False

            # 날짜 필터 적용
            if date_filter:
                filter_date = datetime.strptime(date_filter, "%Y-%m-%d").date()
                if file_date < filter_date:
                    return False

            return True
        except (OSError, ValueError):
            return False

    def has_valid_extension(self, file_path: Path) -> bool:
        """파일 확장자가 유효한지 확인"""
        ext = file_path.suffix.lower()
        valid_exts = self.config['validation']['file_extensions']
        return ext in valid_exts

    def scan_directory(self, equipment_id: str, log_callback, max_depth: int = 2) -> Set[str]:
        """
        장비 디렉토리 스캔 및 차트번호 추출 (최적화 버전)

        Args:
            equipment_id: 장비 ID (SP, TOPO, OCT 등)
            log_callback: 로그 출력 콜백 함수
            max_depth: 최대 탐색 깊이 (기본값: 2)

        Returns:
            차트번호 집합
        """
        equipment = self.config['equipment'][equipment_id]
        path = Path(equipment['path'])
        pattern = equipment['pattern']
        scan_type = equipment['scan_type']
        date_filter = equipment.get('date_filter')

        chart_numbers = set()

        if not path.exists():
            log_callback(f"  ⚠️  경로 없음: {path}")
            return chart_numbers

        try:
            today = date.today()

            def scan_recursive(current_path: Path, depth: int = 0):
                """재귀적으로 디렉토리 스캔 (깊이 제한)"""
                if depth > max_depth:
                    return

                try:
                    items = list(current_path.iterdir())
                except (PermissionError, OSError):
                    return

                for item in items:
                    try:
                        # 날짜 체크를 먼저 해서 빠르게 필터링
                        ctime = os.path.getctime(item)
                        file_date = date.fromtimestamp(ctime)

                        if file_date != today:
                            continue

                        # 날짜 필터 적용
                        if date_filter:
                            filter_date = datetime.strptime(date_filter, "%Y-%m-%d").date()
                            if file_date < filter_date:
                                continue

                        # 파일 처리
                        if item.is_file() and scan_type in ['file', 'both']:
                            # 확장자 체크
                            if not self.has_valid_extension(item):
                                continue

                            # 차트번호 추출
                            filename = item.name
                            match = re.search(pattern, filename)
                            if match:
                                chart_num = match.group(1)
                                if self.is_valid_chart_number(chart_num):
                                    chart_numbers.add(chart_num)

                        # 폴더 처리
                        elif item.is_dir():
                            if scan_type == 'both':
                                # OCT의 경우 폴더명에서도 차트번호 추출
                                folder_name = item.name
                                match = re.search(pattern, folder_name)
                                if match:
                                    chart_num = match.group(1)
                                    if self.is_valid_chart_number(chart_num):
                                        chart_numbers.add(chart_num)

                            # 하위 디렉토리 탐색
                            scan_recursive(item, depth + 1)

                    except (OSError, ValueError):
                        continue

            # 스캔 시작
            scan_recursive(path)

        except Exception as e:
            log_callback(f"  ❌ 오류: {equipment['name']} - {str(e)}")

        return chart_numbers

    def calculate_glaucoma(self, log_callback) -> int:
        """
        녹내장 계산 (HFA ∩ OCT)

        Args:
            log_callback: 로그 출력 콜백 함수

        Returns:
            녹내장 환자 수
        """
        try:
            hfa_charts = self.chart_numbers.get('HFA', set())
            oct_charts = self.chart_numbers.get('OCT', set())
            glaucoma_charts = hfa_charts & oct_charts
            return len(glaucoma_charts)
        except Exception as e:
            log_callback(f"  ❌ 녹내장 계산 오류: {str(e)}")
            return 0

    def calculate_lasik(self, log_callback) -> int:
        """
        라식 계산 ((ORB ∩ TOPO) + SCR 폴더)

        Args:
            log_callback: 로그 출력 콜백 함수

        Returns:
            라식 환자 수
        """
        try:
            orb_charts = self.chart_numbers.get('ORB', set())
            topo_charts = self.chart_numbers.get('TOPO', set())
            lasik_charts = orb_charts & topo_charts

            # SCR 폴더 추가
            scr_path = Path(self.config['special_items']['라식']['scr_folder'])
            if scr_path.exists():
                for item in scr_path.iterdir():
                    if self.is_today_file(item):
                        # 차트번호 추출 시도 (여러 패턴 적용 가능)
                        item_name = item.name
                        for pattern in [r'\s(\d+)\s', r'_(\d+)\.', r'(\d+)']:
                            match = re.search(pattern, item_name)
                            if match:
                                chart_num = match.group(1)
                                if self.is_valid_chart_number(chart_num):
                                    lasik_charts.add(chart_num)
                                    break

            return len(lasik_charts)
        except Exception as e:
            log_callback(f"  ❌ 라식 계산 오류: {str(e)}")
            return 0

    def calculate_fundus(self, log_callback) -> int:
        """
        안저 계산 (FUNDERS + OPTOS 폴더)

        Args:
            log_callback: 로그 출력 콜백 함수

        Returns:
            안저 촬영 환자 수
        """
        fundus_charts = set()
        pattern = self.config['special_items']['안저']['pattern']

        try:
            for folder_str in self.config['special_items']['안저']['folders']:
                folder_path = Path(folder_str)

                if not folder_path.exists():
                    log_callback(f"  ⚠️  경로 없음: {folder_path}")
                    continue

                for item in folder_path.iterdir():
                    if not self.is_today_file(item):
                        continue

                    item_name = item.name
                    match = re.search(pattern, item_name)
                    if match:
                        chart_num = match.group(1)
                        if self.is_valid_chart_number(chart_num):
                            fundus_charts.add(chart_num)

        except Exception as e:
            log_callback(f"  ❌ 안저 계산 오류: {str(e)}")

        return len(fundus_charts)

    def process_reservation_file(self, file_path: str, log_callback) -> Dict[str, int]:
        """
        예약 파일 처리

        Args:
            file_path: 예약 엑셀 파일 경로
            log_callback: 로그 출력 콜백 함수

        Returns:
            키워드별 카운트 딕셔너리
        """
        counts = {'verion': 0, 'lensx': 0, 'ex500': 0}
        found_cells = set()  # 중복 방지

        try:
            # 엑셀 파일 로드
            wb = load_workbook(file_path, data_only=True)

            for sheet in wb.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is None:
                            continue

                        cell_value = str(cell.value).lower()

                        # "수술방법:" 키워드가 있는 셀만 처리
                        if "수술방법:" not in cell_value:
                            continue

                        # 중복 체크 (같은 내용)
                        cell_key = f"{sheet.title}_{cell.coordinate}_{cell_value}"
                        if cell_key in found_cells:
                            continue
                        found_cells.add(cell_key)

                        # Verion (Toric) 키워드
                        if any(kw in cell_value for kw in self.config['reservation']['verion_keywords']):
                            counts['verion'] += 1

                        # Lensx 키워드
                        elif any(kw in cell_value for kw in self.config['reservation']['lensx_keywords']):
                            counts['lensx'] += 1

                        # EX500 키워드
                        elif any(kw in cell_value for kw in self.config['reservation']['ex500_keywords']):
                            counts['ex500'] += 1

            wb.close()

        except Exception as e:
            log_callback(f"  ❌ 예약 파일 처리 오류: {str(e)}")

        return counts

    def write_excel(self, output_path: str, staff_selected: List[str],
                   manual_fag: int, manual_glasses: int,
                   reservation_counts: Dict[str, int], log_callback) -> bool:
        """
        엑셀 파일 작성

        Args:
            output_path: 출력 파일 경로
            staff_selected: 선택된 직원 목록
            manual_fag: FAG 건수
            manual_glasses: 안경검사 건수
            reservation_counts: 예약 파일 카운트
            log_callback: 로그 출력 콜백 함수

        Returns:
            성공 여부
        """
        try:
            # 템플릿 파일 복사
            template_file = self.config['template_file']
            if not os.path.exists(template_file):
                log_callback(f"  ❌ 템플릿 파일 없음: {template_file}")
                return False

            wb = load_workbook(template_file)
            ws = wb[self.config['target_sheet']]

            # 날짜 기입
            date_cell = self.config['date_cell']
            ws.cell(date_cell['row'], date_cell['col']).value = date.today().strftime('%Y-%m-%d')

            # 근무 인원 기입
            staff_cell = self.config['staff_cell']
            staff_count = len(staff_selected)
            staff_text = f"{staff_count}명( {', '.join(staff_selected)} )"
            ws.cell(staff_cell['row'], staff_cell['col']).value = staff_text

            # 각 장비별 결과 기입
            for equipment_id, chart_set in self.chart_numbers.items():
                if equipment_id in self.config['equipment']:
                    cell_info = self.config['equipment'][equipment_id]['cell']
                    ws.cell(cell_info['row'], cell_info['col']).value = len(chart_set)

            # 특수 항목 기입
            # 녹내장
            glaucoma_count = self.calculate_glaucoma(log_callback)
            glaucoma_cell = self.config['special_items']['녹내장']['cell']
            ws.cell(glaucoma_cell['row'], glaucoma_cell['col']).value = glaucoma_count

            # 라식
            lasik_count = self.calculate_lasik(log_callback)
            lasik_cell = self.config['special_items']['라식']['cell']
            ws.cell(lasik_cell['row'], lasik_cell['col']).value = lasik_count

            # 안저
            fundus_count = self.calculate_fundus(log_callback)
            fundus_cell = self.config['special_items']['안저']['cell']
            ws.cell(fundus_cell['row'], fundus_cell['col']).value = fundus_count

            # 수기 입력 항목
            fag_cell = self.config['manual_input']['FAG']
            ws.cell(fag_cell['row'], fag_cell['col']).value = manual_fag

            glasses_cell = self.config['manual_input']['안경검사']
            ws.cell(glasses_cell['row'], glasses_cell['col']).value = manual_glasses

            # 예약 파일 결과 기입
            # Verion (IOL700 장비 결과와 예약 결과 중 큰 값 사용)
            iol700_count = len(self.chart_numbers.get('IOL700', set()))
            verion_count = max(iol700_count, reservation_counts.get('verion', 0))
            verion_cell = self.config['reservation']['cells']['verion']
            ws.cell(verion_cell['row'], verion_cell['col']).value = verion_count

            # Lensx
            lensx_cell = self.config['reservation']['cells']['lensx']
            ws.cell(lensx_cell['row'], lensx_cell['col']).value = reservation_counts.get('lensx', 0)

            # EX500
            ex500_cell = self.config['reservation']['cells']['ex500']
            ws.cell(ex500_cell['row'], ex500_cell['col']).value = reservation_counts.get('ex500', 0)

            # 저장
            wb.save(output_path)
            wb.close()

            log_callback("  ✓ 엑셀 작성 완료")
            return True

        except Exception as e:
            log_callback(f"  ❌ 엑셀 작성 오류: {str(e)}")
            return False

    def convert_to_pdf(self, excel_path: str, pdf_path: str, log_callback) -> bool:
        """
        엑셀 파일을 PDF로 변환

        Args:
            excel_path: 엑셀 파일 경로
            pdf_path: PDF 파일 경로
            log_callback: 로그 출력 콜백 함수

        Returns:
            성공 여부
        """
        if not HAS_WIN32:
            log_callback("  ⚠️  pywin32가 없어 PDF 변환 불가")
            return False

        try:
            # PDF 출력 디렉토리 생성
            os.makedirs(os.path.dirname(pdf_path), exist_ok=True)

            # Excel 실행
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False

            # 워크북 열기
            wb = excel.Workbooks.Open(os.path.abspath(excel_path))
            ws = wb.Worksheets(self.config['target_sheet'])

            # PDF로 저장
            ws.ExportAsFixedFormat(0, os.path.abspath(pdf_path))

            # 정리
            wb.Close(SaveChanges=False)
            excel.Quit()

            log_callback(f"  ✓ PDF 생성 완료: {pdf_path}")
            return True

        except Exception as e:
            log_callback(f"  ❌ PDF 변환 오류: {str(e)}")
            return False


class DailyReportGUI:
    """일일결산 시스템의 GUI 클래스"""

    def __init__(self, root: tk.Tk, system: DailyReportSystem):
        """
        GUI 초기화

        Args:
            root: tkinter 루트 윈도우
            system: DailyReportSystem 인스턴스
        """
        self.root = root
        self.system = system
        self.reservation_files = []

        self.setup_gui()

    def setup_gui(self):
        """GUI 구성 요소 생성"""
        self.root.title("일일결산 자동화 시스템")
        self.root.geometry("900x700")
        self.root.resizable(True, True)

        # 메인 프레임 (좌우 분할)
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(0, weight=1)

        # 좌측 입력 영역
        left_frame = ttk.Frame(main_frame, padding="5")
        left_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # 우측 로그 영역
        right_frame = ttk.Frame(main_frame, padding="5")
        right_frame.grid(row=0, column=1, sticky=(tk.W, tk.E, tk.N, tk.S))
        right_frame.columnconfigure(0, weight=1)
        right_frame.rowconfigure(1, weight=1)

        # === 좌측 영역 구성 ===

        # 1. 근무 인원 선택
        staff_label = ttk.Label(left_frame, text="📋 근무 인원", font=("", 12, "bold"))
        staff_label.grid(row=0, column=0, columnspan=2, sticky=tk.W, pady=(0, 5))

        # 스크롤 가능한 체크박스 프레임
        staff_canvas = tk.Canvas(left_frame, height=200)
        staff_scrollbar = ttk.Scrollbar(left_frame, orient="vertical", command=staff_canvas.yview)
        staff_scrollable = ttk.Frame(staff_canvas)

        staff_scrollable.bind(
            "<Configure>",
            lambda e: staff_canvas.configure(scrollregion=staff_canvas.bbox("all"))
        )

        staff_canvas.create_window((0, 0), window=staff_scrollable, anchor="nw")
        staff_canvas.configure(yscrollcommand=staff_scrollbar.set)

        staff_canvas.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        staff_scrollbar.grid(row=1, column=1, sticky=(tk.N, tk.S), pady=(0, 10))

        # 직원 체크박스 생성
        self.staff_vars = {}
        for i, staff_name in enumerate(self.system.config['staff_list']):
            var = tk.BooleanVar(value=True)  # 기본값: 체크됨
            self.staff_vars[staff_name] = var
            cb = ttk.Checkbutton(staff_scrollable, text=staff_name, variable=var)
            cb.grid(row=i, column=0, sticky=tk.W, padx=5, pady=2)

        # 2. 예약 파일 선택
        ttk.Separator(left_frame, orient='horizontal').grid(row=2, column=0, columnspan=2,
                                                             sticky=(tk.W, tk.E), pady=10)

        reservation_label = ttk.Label(left_frame, text="📁 예약 파일", font=("", 12, "bold"))
        reservation_label.grid(row=3, column=0, columnspan=2, sticky=tk.W, pady=(0, 5))

        self.file_button = ttk.Button(left_frame, text="파일 선택...", command=self.select_files)
        self.file_button.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 5))

        self.file_label = ttk.Label(left_frame, text="선택된 파일: 없음", foreground="gray")
        self.file_label.grid(row=5, column=0, columnspan=2, sticky=tk.W)

        # 3. 수기 입력
        ttk.Separator(left_frame, orient='horizontal').grid(row=6, column=0, columnspan=2,
                                                             sticky=(tk.W, tk.E), pady=10)

        manual_label = ttk.Label(left_frame, text="✍ 수기 입력", font=("", 12, "bold"))
        manual_label.grid(row=7, column=0, columnspan=2, sticky=tk.W, pady=(0, 10))

        # FAG
        fag_label = ttk.Label(left_frame, text="FAG:")
        fag_label.grid(row=8, column=0, sticky=tk.W, padx=(0, 5))

        self.fag_entry = ttk.Entry(left_frame, width=10)
        self.fag_entry.insert(0, "0")
        self.fag_entry.grid(row=8, column=1, sticky=tk.W, pady=3)

        # 안경검사
        glasses_label = ttk.Label(left_frame, text="안경검사:")
        glasses_label.grid(row=9, column=0, sticky=tk.W, padx=(0, 5))

        self.glasses_entry = ttk.Entry(left_frame, width=10)
        self.glasses_entry.insert(0, "0")
        self.glasses_entry.grid(row=9, column=1, sticky=tk.W, pady=3)

        # 4. 실행 버튼
        ttk.Separator(left_frame, orient='horizontal').grid(row=10, column=0, columnspan=2,
                                                             sticky=(tk.W, tk.E), pady=15)

        self.run_button = ttk.Button(left_frame, text="🚀 결산 실행", command=self.run_report)
        self.run_button.grid(row=11, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=10)

        # === 우측 영역 구성 ===

        log_label = ttk.Label(right_frame, text="실행 로그", font=("", 12, "bold"))
        log_label.grid(row=0, column=0, sticky=tk.W, pady=(0, 5))

        self.log_text = scrolledtext.ScrolledText(right_frame, width=50, height=30,
                                                   state='disabled', wrap=tk.WORD)
        self.log_text.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

    def log(self, message: str):
        """로그 메시지 출력"""
        self.log_text.configure(state='normal')
        self.log_text.insert(tk.END, message + '\n')
        self.log_text.see(tk.END)
        self.log_text.configure(state='disabled')
        self.root.update()

    def select_files(self):
        """예약 파일 선택"""
        files = filedialog.askopenfilenames(
            title="예약 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )

        if files:
            self.reservation_files = list(files)
            file_count = len(files)
            self.file_label.config(text=f"선택된 파일: {file_count}개", foreground="blue")
        else:
            self.reservation_files = []
            self.file_label.config(text="선택된 파일: 없음", foreground="gray")

    def get_selected_staff(self) -> List[str]:
        """선택된 직원 목록 반환"""
        return [name for name, var in self.staff_vars.items() if var.get()]

    def run_report(self):
        """결산 실행 (별도 스레드에서)"""
        # 버튼 비활성화
        self.run_button.config(state='disabled')
        self.file_button.config(state='disabled')

        # 로그 초기화
        self.log_text.configure(state='normal')
        self.log_text.delete(1.0, tk.END)
        self.log_text.configure(state='disabled')

        # 스레드 실행
        thread = threading.Thread(target=self.process_report, daemon=True)
        thread.start()

    def process_report(self):
        """결산 처리 메인 로직"""
        try:
            self.log("=" * 54)
            self.log(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 결산 시작")
            self.log("=" * 54)
            self.log("")

            # 1. 디렉토리 자동 스캔
            self.log("[1/4] 디렉토리 자동 스캔 중...")
            for equipment_id in self.system.config['equipment'].keys():
                equipment_name = self.system.config['equipment'][equipment_id]['name']
                self.log(f"  🔍 {equipment_name} 스캔 중...")

                chart_set = self.system.scan_directory(equipment_id, self.log)
                self.system.chart_numbers[equipment_id] = chart_set

                self.log(f"  ✓ {equipment_name}: {len(chart_set)}건")

            self.log("")

            # 2. 특수 항목 계산
            self.log("[특수 항목 계산]")

            glaucoma_count = self.system.calculate_glaucoma(self.log)
            self.log(f"  ✓ 녹내장 (HFA ∩ OCT): {glaucoma_count}건")

            lasik_count = self.system.calculate_lasik(self.log)
            self.log(f"  ✓ 라식 (ORB ∩ TOPO): {lasik_count}건")

            fundus_count = self.system.calculate_fundus(self.log)
            self.log(f"  ✓ 안저: {fundus_count}건")

            self.log("")

            # 3. 예약 파일 처리
            reservation_counts = {'verion': 0, 'lensx': 0, 'ex500': 0}

            if self.reservation_files:
                self.log(f"[2/4] 예약 파일 분석 중... ({len(self.reservation_files)}개 파일)")

                for file_path in self.reservation_files:
                    file_name = os.path.basename(file_path)
                    self.log(f"  📄 {file_name}")

                    file_counts = self.system.process_reservation_file(file_path, self.log)

                    # 누적
                    for key in reservation_counts:
                        reservation_counts[key] += file_counts[key]

                self.log(f"  ✓ Verion (예약): {reservation_counts['verion']}건")
                self.log(f"  ✓ Lensx: {reservation_counts['lensx']}건")
                self.log(f"  ✓ EX500: {reservation_counts['ex500']}건")
            else:
                self.log("[2/4] 예약 파일 선택 안 함 (건너뜀)")

            self.log("")

            # 4. 엑셀 작성
            self.log("[3/4] 엑셀 파일 작성 중...")

            # 선택된 직원
            staff_selected = self.get_selected_staff()
            if not staff_selected:
                self.log("  ⚠️  경고: 직원이 선택되지 않았습니다.")

            # 수기 입력 값
            try:
                manual_fag = int(self.fag_entry.get())
            except ValueError:
                manual_fag = 0
                self.log("  ⚠️  FAG 값이 올바르지 않아 0으로 설정합니다.")

            try:
                manual_glasses = int(self.glasses_entry.get())
            except ValueError:
                manual_glasses = 0
                self.log("  ⚠️  안경검사 값이 올바르지 않아 0으로 설정합니다.")

            # 임시 엑셀 파일 경로
            today_str = date.today().strftime('%Y%m%d')
            temp_excel = f"일일결산_{today_str}_temp.xlsx"

            success = self.system.write_excel(
                temp_excel, staff_selected, manual_fag, manual_glasses,
                reservation_counts, self.log
            )

            if not success:
                self.log("")
                self.log("=" * 54)
                self.log("❌ 결산 실패: 엑셀 작성 오류")
                self.log("=" * 54)
                return

            self.log("")

            # 5. PDF 변환
            self.log("[4/4] PDF 생성 중...")

            pdf_path = self.system.config['output_pdf'].replace('{date}', today_str)
            pdf_success = self.system.convert_to_pdf(temp_excel, pdf_path, self.log)

            self.log("")
            self.log("=" * 54)
            self.log("✅ 결산 완료!")
            self.log("=" * 54)
            self.log("")

            # PDF 열기
            if pdf_success and os.path.exists(pdf_path):
                self.log("📄 PDF 파일을 엽니다...")
                if sys.platform == 'win32':
                    os.startfile(pdf_path)
                else:
                    self.log(f"  PDF 경로: {pdf_path}")

                # 임시 엑셀 파일 삭제
                try:
                    os.remove(temp_excel)
                except:
                    pass
            else:
                self.log(f"📄 엑셀 파일이 저장되었습니다: {temp_excel}")

        except Exception as e:
            self.log("")
            self.log("=" * 54)
            self.log(f"❌ 오류 발생: {str(e)}")
            self.log("=" * 54)

        finally:
            # 버튼 다시 활성화
            self.run_button.config(state='normal')
            self.file_button.config(state='normal')


def main():
    """메인 함수"""
    # 설정 파일 확인
    config_path = "config.json"
    if not os.path.exists(config_path):
        messagebox.showerror("오류", "config.json 파일을 찾을 수 없습니다.")
        sys.exit(1)

    # 시스템 초기화
    system = DailyReportSystem(config_path)

    # GUI 실행
    root = tk.Tk()
    app = DailyReportGUI(root, system)
    root.mainloop()


if __name__ == "__main__":
    main()
