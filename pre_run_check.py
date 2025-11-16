#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
일일결산 프로그램 실행 전 사전 체크 도구

실행 순서:
1. 디렉토리 접근성 확인 및 파일명 샘플 수집
2. 예약 엑셀 파일 읽기 테스트
3. 모든 경로 최종 검증

결과는 텍스트 파일로 저장되어 분석에 사용됩니다.
"""

import os
import sys
import json
from datetime import date, datetime
from pathlib import Path

# 결과 저장 파일
SAMPLE_FILE = "directory_samples.txt"
REPORT_FILE = f"pre_run_report_{date.today().strftime('%Y%m%d')}.txt"


def log_print(message, file_handle=None):
    """화면과 파일에 동시 출력"""
    print(message)
    if file_handle:
        file_handle.write(message + "\n")


def load_config(config_path="config_real.json"):
    """설정 파일 로드"""
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"❌ 설정 파일 없음: {config_path}")
        sys.exit(1)


def guess_today_folder(base_path, equipment_id, folder_structure):
    """오늘 날짜 폴더 경로 생성"""
    if not folder_structure:
        return base_path

    today = date.today()
    year = today.strftime('%Y')
    month = today.strftime('%m')
    day = today.strftime('%d')

    # folder_structure 패턴 치환
    folder_path = folder_structure
    folder_path = folder_path.replace('YYYY', year)
    folder_path = folder_path.replace('MM', month)
    folder_path = folder_path.replace('DD', day)

    return os.path.join(base_path, folder_path)


def collect_file_samples(equipment_id, equipment_info, sample_file, max_samples=50):
    """
    1단계: 디렉토리 파일명 샘플 수집
    실제 파일명들을 텍스트 파일에 기록
    """
    sample_file.write("\n" + "="*100 + "\n")
    sample_file.write(f"장비: {equipment_info['name']} ({equipment_id})\n")
    sample_file.write("="*100 + "\n\n")

    base_path = equipment_info['path']

    # 경로 존재 확인
    if not os.path.exists(base_path):
        sample_file.write(f"❌ 경로 접근 불가: {base_path}\n")
        sample_file.write(f"   → 네트워크 연결 또는 경로 설정 확인 필요\n\n")
        return False

    sample_file.write(f"✅ 기본 경로: {base_path}\n")

    # 오늘 폴더 경로 생성
    folder_structure = equipment_info.get('folder_structure')
    if folder_structure:
        today_folder = guess_today_folder(base_path, equipment_id, folder_structure)
        sample_file.write(f"📂 오늘 폴더 경로: {today_folder}\n")

        if not os.path.exists(today_folder):
            sample_file.write(f"⚠️  오늘 폴더 없음 (기본 경로 사용)\n")
            today_folder = base_path
        else:
            sample_file.write(f"✅ 오늘 폴더 접근 가능\n")
    else:
        today_folder = base_path
        sample_file.write(f"📂 단일 폴더 구조\n")

    sample_file.write(f"\n현재 패턴: {equipment_info.get('pattern', 'N/A')}\n")
    sample_file.write(f"생성일 확인: {equipment_info.get('use_creation_time', False)}\n\n")

    # 파일 샘플 수집
    sample_file.write(f"📄 파일명 샘플 (최대 {max_samples}개):\n")
    sample_file.write("-"*100 + "\n")

    try:
        file_count = 0
        today_date = date.today()

        for item in os.listdir(today_folder):
            item_path = os.path.join(today_folder, item)

            if os.path.isfile(item_path):
                # 파일 생성일 확인
                try:
                    ctime = os.path.getctime(item_path)
                    file_date = date.fromtimestamp(ctime)
                    is_today = file_date == today_date
                    date_str = file_date.strftime('%Y-%m-%d')
                except:
                    is_today = False
                    date_str = "Unknown"

                # 오늘 파일 우선, 아니면 모든 파일
                marker = "🟢" if is_today else "⚪"
                sample_file.write(f"{marker} {item} (생성일: {date_str})\n")

                file_count += 1
                if file_count >= max_samples:
                    break

        if file_count == 0:
            sample_file.write("   (파일 없음)\n")
        else:
            sample_file.write(f"\n총 {file_count}개 샘플 수집 (🟢 = 오늘 생성 파일)\n")

    except Exception as e:
        sample_file.write(f"❌ 파일 읽기 오류: {e}\n")
        return False

    sample_file.write("\n")
    return True


def test_reservation_file(report_file):
    """
    2단계: 예약 엑셀 파일 읽기 테스트
    """
    report_file.write("\n" + "="*100 + "\n")
    report_file.write("2️⃣ 예약 엑셀 파일 읽기 테스트\n")
    report_file.write("="*100 + "\n\n")

    # .xls 파일 지원 확인
    try:
        import xlrd
        report_file.write("✅ xlrd 라이브러리 설치됨 (.xls 파일 지원)\n")
        has_xlrd = True
    except ImportError:
        report_file.write("⚠️  xlrd 라이브러리 미설치 (.xls 파일 읽기 불가)\n")
        report_file.write("   → 설치: pip install xlrd\n")
        has_xlrd = False

    # .xlsx 파일 지원 확인
    try:
        import openpyxl
        report_file.write("✅ openpyxl 라이브러리 설치됨 (.xlsx 파일 지원)\n\n")
        has_openpyxl = True
    except ImportError:
        report_file.write("⚠️  openpyxl 라이브러리 미설치 (.xlsx 파일 읽기 불가)\n")
        report_file.write("   → 설치: pip install openpyxl\n\n")
        has_openpyxl = False

    # 예약 파일 찾기 (일반적인 위치)
    possible_paths = [
        "D:\\결산\\예약.xls",
        "D:\\결산\\예약.xlsx",
        "예약.xls",
        "예약.xlsx",
    ]

    found_file = None
    for path in possible_paths:
        if os.path.exists(path):
            found_file = path
            break

    if not found_file:
        report_file.write("⚠️  예약 파일을 찾을 수 없습니다.\n")
        report_file.write("   → 프로그램 실행 시 직접 파일을 선택하여 테스트하세요.\n")
        report_file.write(f"   확인한 경로: {', '.join(possible_paths)}\n\n")
        return False

    report_file.write(f"✅ 예약 파일 발견: {found_file}\n\n")

    # 파일 읽기 테스트
    try:
        file_ext = os.path.splitext(found_file)[1].lower()

        if file_ext == '.xlsx':
            if not has_openpyxl:
                report_file.write("❌ .xlsx 파일이지만 openpyxl이 없어 읽을 수 없습니다.\n\n")
                return False

            import openpyxl
            wb = openpyxl.load_workbook(found_file, data_only=True)
            ws = wb.active
            total_rows = ws.max_row

            report_file.write(f"📊 엑셀 정보:\n")
            report_file.write(f"   - 형식: .xlsx\n")
            report_file.write(f"   - 시트: {wb.sheetnames}\n")
            report_file.write(f"   - 행 개수: {total_rows}\n\n")

            # 샘플 데이터 읽기
            report_file.write("📋 샘플 데이터 (첫 10행):\n")
            report_file.write("-"*100 + "\n")

            for row_idx in range(1, min(11, total_rows + 1)):
                row_data = []
                for col_idx in range(1, min(11, ws.max_column + 1)):
                    cell_value = ws.cell(row_idx, col_idx).value
                    if cell_value:
                        row_data.append(str(cell_value))
                if row_data:
                    report_file.write(f"행 {row_idx}: {' | '.join(row_data)}\n")

        elif file_ext == '.xls':
            if not has_xlrd:
                report_file.write("❌ .xls 파일이지만 xlrd가 없어 읽을 수 없습니다.\n\n")
                return False

            import xlrd
            wb = xlrd.open_workbook(found_file)
            ws = wb.sheet_by_index(0)
            total_rows = ws.nrows

            report_file.write(f"📊 엑셀 정보:\n")
            report_file.write(f"   - 형식: .xls\n")
            report_file.write(f"   - 시트: {wb.sheet_names()}\n")
            report_file.write(f"   - 행 개수: {total_rows}\n\n")

            # 샘플 데이터 읽기
            report_file.write("📋 샘플 데이터 (첫 10행):\n")
            report_file.write("-"*100 + "\n")

            for row_idx in range(min(10, total_rows)):
                row_data = []
                for col_idx in range(min(10, ws.ncols)):
                    cell_value = ws.cell_value(row_idx, col_idx)
                    if cell_value:
                        row_data.append(str(cell_value))
                if row_data:
                    report_file.write(f"행 {row_idx + 1}: {' | '.join(row_data)}\n")

        report_file.write("\n✅ 예약 파일 읽기 성공\n\n")
        return True

    except Exception as e:
        report_file.write(f"❌ 예약 파일 읽기 실패: {e}\n\n")
        return False


def verify_all_paths(config, report_file):
    """
    3단계: 모든 디렉토리 경로 검증
    """
    report_file.write("\n" + "="*100 + "\n")
    report_file.write("3️⃣ 디렉토리 경로 최종 검증\n")
    report_file.write("="*100 + "\n\n")

    all_ok = True

    # 장비 경로 검증
    report_file.write("📁 장비 경로:\n")
    report_file.write("-"*100 + "\n")

    for equipment_id, equipment_info in config['equipment'].items():
        base_path = equipment_info['path']
        name = equipment_info['name']

        if os.path.exists(base_path):
            # 오늘 폴더 확인
            folder_structure = equipment_info.get('folder_structure')
            if folder_structure:
                today_folder = guess_today_folder(base_path, equipment_id, folder_structure)
                if os.path.exists(today_folder):
                    report_file.write(f"✅ {name:10s} ({equipment_id:6s}): {today_folder}\n")
                else:
                    report_file.write(f"⚠️  {name:10s} ({equipment_id:6s}): 오늘 폴더 없음 ({today_folder})\n")
                    all_ok = False
            else:
                report_file.write(f"✅ {name:10s} ({equipment_id:6s}): {base_path}\n")
        else:
            report_file.write(f"❌ {name:10s} ({equipment_id:6s}): 접근 불가 ({base_path})\n")
            all_ok = False

    # 특별 항목 경로 검증 (안저)
    report_file.write("\n📁 특별 항목 경로:\n")
    report_file.write("-"*100 + "\n")

    if '안저' in config['special_items']:
        fundus_info = config['special_items']['안저']
        for idx, folder_path in enumerate(fundus_info.get('folders', []), 1):
            if '[TODO' in folder_path or '[하위폴더]' in folder_path:
                report_file.write(f"⚠️  안저 경로 {idx}: 미설정 ({folder_path})\n")
                report_file.write(f"   → find_fundus_path.py 실행하여 정확한 경로 확인 필요\n")
                all_ok = False
            elif os.path.exists(folder_path):
                report_file.write(f"✅ 안저 경로 {idx}: {folder_path}\n")
            else:
                report_file.write(f"❌ 안저 경로 {idx}: 접근 불가 ({folder_path})\n")
                all_ok = False

    report_file.write("\n")

    if all_ok:
        report_file.write("✅ 모든 경로 검증 완료\n\n")
    else:
        report_file.write("⚠️  일부 경로에 문제가 있습니다. 위 내용을 확인하세요.\n\n")

    return all_ok


def main():
    """메인 실행"""
    print("\n" + "="*100)
    print("🔍 일일결산 프로그램 사전 체크 시작")
    print("="*100 + "\n")

    # 설정 파일 로드
    config = load_config()

    # 결과 파일 오픈
    sample_file = open(SAMPLE_FILE, 'w', encoding='utf-8')
    report_file = open(REPORT_FILE, 'w', encoding='utf-8')

    try:
        # 헤더 작성
        sample_file.write("="*100 + "\n")
        sample_file.write("일일결산 시스템 - 디렉토리 파일명 샘플 수집\n")
        sample_file.write(f"수집 날짜: {date.today().strftime('%Y년 %m월 %d일')}\n")
        sample_file.write("="*100 + "\n")
        sample_file.write("\n이 파일은 각 장비 디렉토리의 실제 파일명 샘플을 포함합니다.\n")
        sample_file.write("파일명 패턴 분석에 사용되며, 올바른 정규식 패턴을 생성하는데 활용됩니다.\n")

        report_file.write("="*100 + "\n")
        report_file.write("일일결산 시스템 - 사전 체크 리포트\n")
        report_file.write(f"체크 날짜: {date.today().strftime('%Y년 %m월 %d일')}\n")
        report_file.write("="*100 + "\n\n")

        # 1단계: 파일명 샘플 수집
        print("\n1️⃣ 디렉토리 파일명 샘플 수집 중...\n")
        report_file.write("1️⃣ 디렉토리 파일명 샘플 수집\n")
        report_file.write("="*100 + "\n\n")

        success_count = 0
        total_count = len(config['equipment'])

        for equipment_id, equipment_info in config['equipment'].items():
            print(f"   처리 중: {equipment_info['name']} ({equipment_id})...")
            if collect_file_samples(equipment_id, equipment_info, sample_file):
                success_count += 1
                report_file.write(f"✅ {equipment_info['name']:10s} ({equipment_id}): 샘플 수집 완료\n")
            else:
                report_file.write(f"❌ {equipment_info['name']:10s} ({equipment_id}): 샘플 수집 실패\n")

        print(f"\n   완료: {success_count}/{total_count}개 장비")
        report_file.write(f"\n수집 완료: {success_count}/{total_count}개 장비\n")
        report_file.write(f"샘플 파일: {SAMPLE_FILE}\n")

        # 2단계: 예약 파일 테스트
        print("\n2️⃣ 예약 엑셀 파일 읽기 테스트 중...\n")
        test_reservation_file(report_file)

        # 3단계: 경로 검증
        print("\n3️⃣ 모든 경로 최종 검증 중...\n")
        all_ok = verify_all_paths(config, report_file)

        # 최종 요약
        report_file.write("\n" + "="*100 + "\n")
        report_file.write("📋 최종 요약\n")
        report_file.write("="*100 + "\n\n")

        report_file.write(f"파일명 샘플: {success_count}/{total_count}개 장비 수집 완료\n")
        report_file.write(f"샘플 파일: {SAMPLE_FILE}\n\n")

        if all_ok:
            report_file.write("✅ 모든 체크 통과 - 프로그램 실행 준비 완료\n\n")
            print("\n✅ 모든 체크 통과!")
            print(f"\n📄 생성된 파일:")
            print(f"   - {SAMPLE_FILE}: 파일명 샘플 (패턴 분석용)")
            print(f"   - {REPORT_FILE}: 상세 리포트")
        else:
            report_file.write("⚠️  일부 문제 발견 - 위 내용을 확인하고 수정 필요\n\n")
            print("\n⚠️  일부 문제가 발견되었습니다.")
            print(f"\n📄 생성된 파일:")
            print(f"   - {SAMPLE_FILE}: 파일명 샘플")
            print(f"   - {REPORT_FILE}: 상세 리포트 (문제점 포함)")

        report_file.write("다음 단계:\n")
        report_file.write("1. directory_samples.txt 파일을 확인하여 파일명 패턴 검토\n")
        report_file.write("2. 문제가 있는 경로는 config_real.json 수정\n")
        report_file.write("3. 안저 경로 미설정 시 find_fundus_path.py 실행\n")
        report_file.write("4. 모든 문제 해결 후 daily_report_fast.py 실행\n")

    finally:
        sample_file.close()
        report_file.close()

    print("\n" + "="*100)
    print("🎉 사전 체크 완료")
    print("="*100 + "\n")


if __name__ == "__main__":
    main()
